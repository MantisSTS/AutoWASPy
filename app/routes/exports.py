"""
Export routes for CSV, Excel, and Markdown formats
"""
import io
import json
import csv
from flask import Blueprint, make_response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app import db
from app.models import Project, TestItem
from app.utils import utc_now

bp = Blueprint('exports', __name__, url_prefix='/project')

@bp.route('/<int:project_id>/export/csv')
def export_csv(project_id):
    """Export project test results to CSV format"""
    project = db.get_or_404(Project, project_id)
    test_items = TestItem.query.filter_by(project_id=project_id).order_by(TestItem.category, TestItem.owasp_id).all()
    
    # Create CSV data
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow([
        'OWASP ID',
        'Title', 
        'Category',
        'Test Type',
        'Finding Status',
        'Risk Level',
        'Evidence',
        'Is Tested',
        'Created Date',
        'Updated Date'
    ])
    
    # Write test items
    for item in test_items:
        writer.writerow([
            item.owasp_id,
            item.title,
            item.category,
            item.test_type.upper(),
            item.finding_status.replace('_', ' ').title(),
            item.risk_level.title() if item.risk_level else 'N/A',
            item.evidence or 'No evidence provided',
            'Yes' if item.is_tested else 'No',
            item.created_date.strftime('%Y-%m-%d %H:%M:%S'),
            item.updated_date.strftime('%Y-%m-%d %H:%M:%S')
        ])
    
    # Create response
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename="{project.name}_security_tests.csv"'
    
    return response

@bp.route('/<int:project_id>/export/xlsx')
def export_xlsx(project_id):
    """Export project test results to Excel format"""
    project = db.get_or_404(Project, project_id)
    test_items = TestItem.query.filter_by(project_id=project_id).order_by(TestItem.category, TestItem.owasp_id).all()
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Security Test Results"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    headers = [
        'OWASP ID', 'Title', 'Category', 'Test Type', 'Finding Status', 
        'Risk Level', 'Evidence', 'Is Tested', 'Created Date', 'Updated Date'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Write data
    for row, item in enumerate(test_items, 2):
        ws.cell(row=row, column=1, value=item.owasp_id)
        ws.cell(row=row, column=2, value=item.title)
        ws.cell(row=row, column=3, value=item.category)
        ws.cell(row=row, column=4, value=item.test_type.upper())
        ws.cell(row=row, column=5, value=item.finding_status.replace('_', ' ').title())
        ws.cell(row=row, column=6, value=item.risk_level.title() if item.risk_level else 'N/A')
        ws.cell(row=row, column=7, value=item.evidence or 'No evidence provided')
        ws.cell(row=row, column=8, value='Yes' if item.is_tested else 'No')
        ws.cell(row=row, column=9, value=item.created_date.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row, column=10, value=item.updated_date.strftime('%Y-%m-%d %H:%M:%S'))
        
        # Color code finding status
        status_cell = ws.cell(row=row, column=5)
        if item.finding_status == 'pass':
            status_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        elif item.finding_status == 'fail':
            status_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        elif item.finding_status == 'informational':
            status_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    
    # Adjust column widths
    column_widths = [15, 50, 30, 12, 15, 12, 60, 12, 20, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename="{project.name}_security_tests.xlsx"'
    
    return response

@bp.route('/<int:project_id>/export/markdown')
def export_markdown(project_id):
    """Export project test results to Markdown format"""
    project = db.get_or_404(Project, project_id)
    test_items = TestItem.query.filter_by(project_id=project_id).order_by(TestItem.category, TestItem.owasp_id).all()
    
    # Group tests by category
    categories = {}
    for item in test_items:
        if item.category not in categories:
            categories[item.category] = []
        categories[item.category].append(item)
    
    # Generate markdown content
    markdown_content = []
    markdown_content.append(f"# {project.name} - Security Test Report")
    markdown_content.append(f"\n**Client:** {project.client_name}")
    markdown_content.append(f"**Test Type:** {project.job_type.replace('_', ' ').title()}")
    markdown_content.append(f"**Generated:** {utc_now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    if project.description:
        markdown_content.append(f"\n**Description:** {project.description}")
    
    if project.urls:
        # Handle both JSON and plain text URL storage
        try:
            urls = json.loads(project.urls) if project.urls.startswith('[') or project.urls.startswith('{') else [url.strip() for url in project.urls.split('\n') if url.strip()]
        except (json.JSONDecodeError, AttributeError):
            urls = [url.strip() for url in project.urls.split('\n') if url.strip()]
        
        if urls:
            markdown_content.append(f"\n**Target URLs:**")
            for url in urls:
                markdown_content.append(f"- {url}")
    
    # Summary statistics
    total_tests = len(test_items)
    tested_items = [item for item in test_items if item.is_tested]
    passed_tests = len([item for item in test_items if item.finding_status == 'pass'])
    failed_tests = len([item for item in test_items if item.finding_status == 'fail'])
    
    markdown_content.append(f"\n## Summary")
    markdown_content.append(f"- **Total Tests:** {total_tests}")
    markdown_content.append(f"- **Tests Completed:** {len(tested_items)}")
    markdown_content.append(f"- **Passed:** {passed_tests}")
    markdown_content.append(f"- **Failed:** {failed_tests}")
    markdown_content.append(f"- **Progress:** {(len(tested_items) / total_tests * 100):.1f}%")
    
    # Test results by category
    markdown_content.append(f"\n## Test Results by Category")
    
    for category, items in categories.items():
        markdown_content.append(f"\n### {category}")
        markdown_content.append(f"\n| OWASP ID | Title | Status | Risk Level | Evidence |")
        markdown_content.append(f"|----------|-------|--------|------------|----------|")
        
        for item in items:
            status_emoji = {
                'pass': '✅',
                'fail': '❌', 
                'informational': 'ℹ️',
                'error': '⚠️',
                'not_tested': '⏳'
            }.get(item.finding_status, '⏳')
            
            risk_level = item.risk_level.title() if item.risk_level else 'N/A'
            evidence = (item.evidence or 'No evidence provided').replace('\n', ' ').replace('|', '\\|')[:100]
            if len(item.evidence or '') > 100:
                evidence = evidence[:97] + '...'
            elif not item.evidence:
                evidence = 'No evidence provided'
            
            markdown_content.append(f"| {item.owasp_id} | {item.title.replace('|', '\\|')} | {status_emoji} {item.finding_status.replace('_', ' ').title()} | {risk_level} | {evidence} |")
    
    # Detailed findings for failed tests
    failed_items = [item for item in test_items if item.finding_status == 'fail']
    if failed_items:
        markdown_content.append(f"\n## Detailed Findings")
        
        for item in failed_items:
            markdown_content.append(f"\n### {item.owasp_id}: {item.title}")
            markdown_content.append(f"**Category:** {item.category}")
            markdown_content.append(f"**Risk Level:** {item.risk_level.title() if item.risk_level else 'Not Specified'}")
            
            if item.evidence:
                markdown_content.append(f"\n**Evidence:**")
                markdown_content.append(f"```")
                markdown_content.append(item.evidence)
                markdown_content.append(f"```")
            
            markdown_content.append(f"\n---")
    
    # Create response
    content = '\n'.join(markdown_content)
    response = make_response(content)
    response.headers['Content-Type'] = 'text/markdown'
    response.headers['Content-Disposition'] = f'attachment; filename="{project.name}_security_report.md"'
    
    return response
